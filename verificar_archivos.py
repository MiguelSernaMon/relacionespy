#!/usr/bin/env python3
"""
Script para verificar la estructura y compatibilidad de los archivos Excel
"""
import pandas as pd
import openpyxl

def leer_excel_inteligente(filepath):
    """Lee un archivo Excel detectando automáticamente dónde comienzan los datos reales"""
    try:
        df = pd.read_excel(filepath)
        columnas_conocidas = ['idOrder', 'authorizationNumber', 'typeOrder', 'identificationPatient', 
                             'nit', 'Nrodcto']
        if any(col in df.columns for col in columnas_conocidas):
            return df, 0
    except:
        pass
    
    # Buscar los encabezados usando openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    fila_encabezados = None
    columnas_objetivo = ['idOrder', 'authorizationNumber', 'identificationPatient', 'nit', 'Nrodcto']
    
    for fila in range(1, min(20, ws.max_row + 1)):
        valores_fila = []
        for columna in range(1, min(50, ws.max_column + 1)):
            celda = ws.cell(row=fila, column=columna)
            if celda.value:
                valores_fila.append(str(celda.value).strip())
        
        coincidencias = sum(1 for col in columnas_objetivo if col in valores_fila)
        if coincidencias >= 1:
            fila_encabezados = fila - 1
            print(f"✅ Encabezados encontrados en fila {fila}")
            break
    
    wb.close()
    
    if fila_encabezados is not None and fila_encabezados > 0:
        df = pd.read_excel(filepath, skiprows=fila_encabezados)
    else:
        df = pd.read_excel(filepath)
    
    return df, fila_encabezados if fila_encabezados else 0

print("="*80)
print("VERIFICACIÓN DE ARCHIVOS PARA LIBRO MEDELLÍN")
print("="*80)

# Archivo 1: Planilla Principal Helpharma (MADRE)
print("\n📁 Archivo 1: Planilla Principal Helpharma.xlsx (PLANILLA MADRE)")
print("-"*80)
try:
    df_madre, skip_rows = leer_excel_inteligente('Planillas Principal Helpharma.xlsx')
    print(f"✅ Archivo leído correctamente (skiprows={skip_rows})")
    print(f"📊 Total de registros: {len(df_madre)}")
    print(f"📋 Columnas encontradas ({len(df_madre.columns)}): {df_madre.columns.tolist()}")
    
    # Verificar columna identificationPatient
    if 'identificationPatient' in df_madre.columns:
        print(f"\n✅ Columna 'identificationPatient' EXISTE")
        df_madre['identificationPatient'] = df_madre['identificationPatient'].astype(str).str.strip()
        nits_madre = df_madre['identificationPatient'].dropna()
        print(f"   - Total de valores: {len(nits_madre)}")
        print(f"   - Valores únicos: {nits_madre.nunique()}")
        print(f"   - Valores vacíos/nulos: {df_madre['identificationPatient'].isna().sum()}")
        print(f"   - Primeros 10 valores:")
        for i, nit in enumerate(nits_madre.head(10), 1):
            print(f"      {i}. '{nit}' (tipo: {type(nit).__name__}, longitud: {len(str(nit))})")
    else:
        print(f"\n❌ Columna 'identificationPatient' NO EXISTE")
        print(f"   Columnas disponibles: {df_madre.columns.tolist()}")
    
    # Verificar otras columnas importantes
    columnas_importantes = ['idOrder', 'authorizationNumber', 'namePatient', 'addressPatient', 
                           'mobilePhonePatient', 'cityNameOrder']
    print(f"\n📋 Otras columnas importantes:")
    for col in columnas_importantes:
        if col in df_madre.columns:
            valores_no_nulos = df_madre[col].dropna()
            print(f"   ✅ {col}: {len(valores_no_nulos)} valores")
        else:
            print(f"   ❌ {col}: NO EXISTE")
    
except Exception as e:
    print(f"❌ Error al leer el archivo: {e}")
    import traceback
    traceback.print_exc()

# Archivo 2: Planilla Inicial 04 Noviembre (OFIMATIC)
print("\n\n📁 Archivo 2: Planillas Iniciales 04 Noviembre 2025.xlsx (PLANILLA OFIMATIC)")
print("-"*80)
try:
    # Intentar leer con diferentes skiprows
    df_ofimatic = None
    for skiprows in [0, 3, 4]:
        try:
            df_test = pd.read_excel('Planillas Iniciales 04 Noviembre 2025.xlsx', skiprows=skiprows)
            if 'nit' in df_test.columns and 'Nrodcto' in df_test.columns:
                df_ofimatic = df_test
                print(f"✅ Archivo leído correctamente (skiprows={skiprows})")
                break
        except:
            continue
    
    if df_ofimatic is None:
        df_ofimatic, skip_rows = leer_excel_inteligente('Planillas Iniciales 04 Noviembre 2025.xlsx')
        print(f"✅ Archivo leído con función inteligente (skiprows={skip_rows})")
    
    print(f"📊 Total de registros: {len(df_ofimatic)}")
    print(f"📋 Columnas encontradas ({len(df_ofimatic.columns)}): {df_ofimatic.columns.tolist()}")
    
    # Verificar columna nit
    if 'nit' in df_ofimatic.columns:
        print(f"\n✅ Columna 'nit' EXISTE")
        df_ofimatic['nit'] = df_ofimatic['nit'].astype(str).str.strip()
        nits_ofimatic = df_ofimatic['nit'].dropna()
        print(f"   - Total de valores: {len(nits_ofimatic)}")
        print(f"   - Valores únicos: {nits_ofimatic.nunique()}")
        print(f"   - Valores vacíos/nulos: {df_ofimatic['nit'].isna().sum()}")
        print(f"   - Primeros 10 valores:")
        for i, nit in enumerate(nits_ofimatic.head(10), 1):
            print(f"      {i}. '{nit}' (tipo: {type(nit).__name__}, longitud: {len(str(nit))})")
    else:
        print(f"\n❌ Columna 'nit' NO EXISTE")
        print(f"   Columnas disponibles: {df_ofimatic.columns.tolist()}")
    
    # Verificar otras columnas importantes
    columnas_importantes = ['Nrodcto', 'NomMensajero', 'NOMBRE', 'DIRECCION', 'TEL1', 'TEL2', 
                           'TipoVta', 'Destino']
    print(f"\n📋 Otras columnas importantes:")
    for col in columnas_importantes:
        if col in df_ofimatic.columns:
            valores_no_nulos = df_ofimatic[col].dropna()
            print(f"   ✅ {col}: {len(valores_no_nulos)} valores")
        else:
            print(f"   ❌ {col}: NO EXISTE")
    
except Exception as e:
    print(f"❌ Error al leer el archivo: {e}")
    import traceback
    traceback.print_exc()

# Comparación de NITs
print("\n\n" + "="*80)
print("ANÁLISIS DE COMPATIBILIDAD")
print("="*80)

try:
    if 'identificationPatient' in df_madre.columns and 'nit' in df_ofimatic.columns:
        # Normalizar NITs
        nits_madre_set = set(df_madre['identificationPatient'].astype(str).str.strip().dropna())
        nits_ofimatic_set = set(df_ofimatic['nit'].astype(str).str.strip().dropna())
        
        # Encontrar NITs en común
        nits_comunes = nits_madre_set.intersection(nits_ofimatic_set)
        
        print(f"\n📊 Estadísticas de NITs:")
        print(f"   - NITs únicos en planilla madre: {len(nits_madre_set)}")
        print(f"   - NITs únicos en planilla ofimatic: {len(nits_ofimatic_set)}")
        print(f"   - NITs EN COMÚN: {len(nits_comunes)}")
        
        if len(nits_comunes) > 0:
            print(f"\n✅ ¡HAY {len(nits_comunes)} NITs EN COMÚN!")
            print(f"   Porcentaje de coincidencia (sobre madre): {len(nits_comunes)/len(nits_madre_set)*100:.1f}%")
            print(f"   Porcentaje de coincidencia (sobre ofimatic): {len(nits_comunes)/len(nits_ofimatic_set)*100:.1f}%")
            print(f"\n   Primeros 15 NITs en común:")
            for i, nit in enumerate(sorted(list(nits_comunes))[:15], 1):
                print(f"      {i}. {nit}")
        else:
            print(f"\n❌ ¡NO HAY NITs EN COMÚN!")
            print(f"\n   Comparación de formatos:")
            print(f"   Ejemplos de NITs en MADRE:")
            for nit in sorted(list(nits_madre_set))[:5]:
                print(f"      - '{nit}' (longitud: {len(nit)})")
            print(f"\n   Ejemplos de NITs en OFIMATIC:")
            for nit in sorted(list(nits_ofimatic_set))[:5]:
                print(f"      - '{nit}' (longitud: {len(nit)})")
            
            # Análisis adicional: buscar similitudes
            print(f"\n   🔍 Buscando similitudes...")
            for nit_madre in sorted(list(nits_madre_set))[:10]:
                for nit_ofimatic in sorted(list(nits_ofimatic_set))[:10]:
                    # Verificar si uno contiene al otro
                    if nit_madre in nit_ofimatic or nit_ofimatic in nit_madre:
                        print(f"      ⚠️ Similitud encontrada: '{nit_madre}' ↔ '{nit_ofimatic}'")
    else:
        print("\n❌ No se pueden comparar NITs porque falta alguna columna")
        if 'identificationPatient' not in df_madre.columns:
            print("   - Falta 'identificationPatient' en planilla madre")
        if 'nit' not in df_ofimatic.columns:
            print("   - Falta 'nit' en planilla ofimatic")

except Exception as e:
    print(f"\n❌ Error al comparar NITs: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "="*80)
print("VERIFICACIÓN COMPLETA")
print("="*80)
