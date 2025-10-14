#!/usr/bin/env python3
"""
Creador de Relaciones Mailbox - Versión Web
Una aplicación web local para procesar planillas CSV y Excel sin problemas de Tkinter
"""

import os
import sys
import webbrowser
import threading
import time
from http.server import HTTPServer, SimpleHTTPRequestHandler
import urllib.parse
import json
import pandas as pd
from io import StringIO, BytesIO
import base64
import openpyxl

def leer_excel_inteligente_desde_contenido(contenido):
    """
    Lee un archivo Excel desde contenido binario detectando automáticamente dónde comienzan los datos reales
    """
    # Primero intentar leer normalmente
    try:
        df = pd.read_excel(BytesIO(contenido))
        # Verificar si tiene columnas conocidas (madre, ofimatic, o ehlpharma)
        columnas_madre = ['idOrder', 'authorizationNumber', 'typeOrder', 'identificationPatient']
        columnas_ofimatic = ['nit', 'Nrodcto']
        columnas_ehlpharma = ['IDENTIFICACION', 'NUMERO DE PEDIDO', 'DIRECCION DE ENTREGA', 'CELULAR', 'DOCUMENTO ASOCIADO']
        
        if any(col in df.columns for col in columnas_madre + columnas_ofimatic + columnas_ehlpharma):
            return df
    except:
        pass
    
    # Si falla, buscar los encabezados usando openpyxl
    wb = openpyxl.load_workbook(BytesIO(contenido), data_only=True)
    ws = wb.active
    
    # Buscar la fila que contiene los encabezados
    fila_encabezados = None
    columnas_objetivo = [
        'idOrder', 'authorizationNumber', 'identificationPatient',  # Medellín
        'nit', 'Nrodcto',  # Ofimatic
        'IDENTIFICACION', 'NUMERO DE PEDIDO', 'DIRECCION DE ENTREGA', 'CELULAR', 'DOCUMENTO ASOCIADO'  # Ehlpharma
    ]
    
    for fila in range(1, min(20, ws.max_row + 1)):  # Buscar en las primeras 20 filas
        valores_fila = []
        for columna in range(1, min(50, ws.max_column + 1)):  # Buscar en las primeras 50 columnas
            celda = ws.cell(row=fila, column=columna)
            if celda.value:
                valores_fila.append(str(celda.value).strip())
        
        # Verificar si esta fila contiene al menos 1 de las columnas objetivo
        coincidencias = sum(1 for col in columnas_objetivo if col in valores_fila)
        if coincidencias >= 1:
            fila_encabezados = fila - 1  # -1 porque skiprows cuenta desde 0
            print(f"✅ Encabezados encontrados en fila {fila}")
            break
    
    wb.close()
    
    # Leer el archivo con skiprows si encontramos los encabezados
    if fila_encabezados is not None and fila_encabezados > 0:
        df = pd.read_excel(BytesIO(contenido), skiprows=fila_encabezados)
    else:
        # Si no encontramos encabezados, intentar con skiprows común
        try:
            df = pd.read_excel(BytesIO(contenido), skiprows=4)
        except:
            df = pd.read_excel(BytesIO(contenido))
    
    return df

def leer_archivo_ofimatic_desde_contenido(contenido, nombre_archivo):
    """
    Lee contenido de archivo ofimatic (CSV o Excel) detectando automáticamente los headers
    """
    try:
        extension = os.path.splitext(nombre_archivo)[1].lower()
        
        if extension == '.csv':
            # Para CSV, decodificar y buscar headers
            codificaciones = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
            
            for encoding in codificaciones:
                try:
                    texto = contenido.decode(encoding)
                    lines = texto.strip().split('\n')
                    
                    # Buscar la fila que contiene 'nit' y 'Nrodcto'
                    for skip_rows in range(min(10, len(lines))):
                        try:
                            contenido_procesado = '\n'.join(lines[skip_rows:])
                            df_test = pd.read_csv(StringIO(contenido_procesado), delimiter=';')
                            if 'nit' in df_test.columns and 'Nrodcto' in df_test.columns:
                                print(f"✅ Headers encontrados en fila {skip_rows + 1} con codificación {encoding}")
                                return df_test
                        except:
                            try:
                                contenido_procesado = '\n'.join(lines[skip_rows:])
                                df_test = pd.read_csv(StringIO(contenido_procesado), delimiter=',')
                                if 'nit' in df_test.columns and 'Nrodcto' in df_test.columns:
                                    print(f"✅ Headers encontrados en fila {skip_rows + 1} con codificación {encoding}")
                                    return df_test
                            except:
                                continue
                except:
                    continue
                    
        elif extension in ['.xlsx', '.xls']:
            # Usar la función inteligente para Excel
            df = leer_excel_inteligente_desde_contenido(contenido)
            
            # Verificar que tenga las columnas necesarias para ofimatic
            if 'nit' not in df.columns or 'Nrodcto' not in df.columns:
                print("⚠️ Columnas 'nit' y 'Nrodcto' no encontradas, intentando detectar automáticamente...")
                
                # Si no tiene las columnas correctas, intentar detectar automáticamente
                for col_idx in range(min(15, len(df.columns))):
                    if df.iloc[:, col_idx].dtype in ['int64', 'float64'] or df.iloc[:, col_idx].astype(str).str.isdigit().sum() > len(df) * 0.7:
                        # Esta parece ser la columna NIT
                        for nrodcto_idx in range(len(df.columns)):
                            if nrodcto_idx != col_idx and len(df.iloc[:, nrodcto_idx].dropna()) > 0:
                                # Renombrar las columnas
                                df_result = df.copy()
                                df_result.columns = [f'col_{i}' for i in range(len(df_result.columns))]
                                df_result = df_result.rename(columns={f'col_{col_idx}': 'nit', f'col_{nrodcto_idx}': 'Nrodcto'})
                                print(f"🔍 Detectadas columnas: nit=columna_{col_idx}, Nrodcto=columna_{nrodcto_idx}")
                                return df_result
                
                raise ValueError("No se pueden identificar las columnas 'nit' y 'Nrodcto'")
            
            return df
        
        raise Exception("No se pudo leer el archivo ofimatic")
        
    except Exception as e:
        raise Exception(f"Error al leer archivo ofimatic {nombre_archivo}: {str(e)}")

def leer_archivo_desde_contenido(contenido, nombre_archivo, es_ofimatic=False):
    """
    Lee contenido de archivo CSV o Excel y retorna un DataFrame
    """
    try:
        # Detectar la extensión del archivo
        extension = os.path.splitext(nombre_archivo)[1].lower()
        
        if extension == '.csv':
            # Para archivos CSV
            if es_ofimatic:
                # Saltar las primeras 4 filas para archivos ofimatic
                lines = contenido.strip().split('\n')
                if len(lines) <= 4:
                    raise ValueError('El archivo ofimatic debe tener más de 4 filas')
                contenido_procesado = '\n'.join(lines[4:])
            else:
                contenido_procesado = contenido
            
            # Intentar leer CSV con diferentes delimitadores
            try:
                df = pd.read_csv(StringIO(contenido_procesado), delimiter=';')
            except:
                try:
                    df = pd.read_csv(StringIO(contenido_procesado), delimiter=',')
                except:
                    df = pd.read_csv(StringIO(contenido_procesado))
            return df
            
        elif extension in ['.xlsx', '.xls']:
            # Para archivos Excel, el contenido viene como bytes
            if isinstance(contenido, str):
                # Si viene como string base64, decodificar
                try:
                    contenido_bytes = base64.b64decode(contenido)
                except:
                    # Si no es base64, convertir a bytes
                    contenido_bytes = contenido.encode('utf-8')
            else:
                contenido_bytes = contenido
            
            # Leer archivo Excel con función inteligente
            df = leer_excel_inteligente_desde_contenido(contenido_bytes)
            return df
        else:
            raise ValueError(f"Formato de archivo no soportado: {extension}. Use CSV, XLS o XLSX")
            
    except Exception as e:
        raise Exception(f"Error al leer el archivo {nombre_archivo}: {str(e)}")


def leer_planilla_inicial_bogota(contenido):
    """
    Lee la planilla inicial de Bogotá manteniendo el formato original.
    La planilla tiene 3 filas de encabezado antes de los datos.
    """
    # Leer todo el archivo sin procesar
    df_completo = pd.read_excel(BytesIO(contenido), header=None)
    
    # Los encabezados están en la fila 3 (índice 3)
    encabezados = df_completo.iloc[3].tolist()
    
    # Los datos comienzan desde la fila 4 (índice 4)
    df_datos = pd.read_excel(BytesIO(contenido), skiprows=4)
    df_datos.columns = encabezados
    
    # Guardar las primeras 4 filas para mantener el formato original
    filas_encabezado = df_completo.iloc[0:4]
    
    return df_datos, filas_encabezado, encabezados


def leer_planilla_pedidos_bogota(contenido):
    """
    Lee la planilla de pedidos con la estructura actual.
    """
    df = pd.read_excel(BytesIO(contenido))
    
    # Convertir IDENTIFICACION a string para facilitar la comparación
    if 'IDENTIFICACION' in df.columns:
        df['IDENTIFICACION'] = df['IDENTIFICACION'].astype(str).str.strip()
    
    return df


def relacionar_por_nit_bogota(df_inicial, df_pedidos):
    """
    Relaciona las planillas por NIT o por DOCUMENTO ASOCIADO.
    Actualiza el campo Nrodcto con el formato: Nrodcto-NUMERO_DE_PEDIDO
    
    Intenta dos métodos de relación:
    1. Por NIT (nit == IDENTIFICACION)
    2. Por documento (Nrodcto normalizado == DOCUMENTO ASOCIADO normalizado)
    """
    # Convertir nit a string para comparación
    df_inicial['nit'] = df_inicial['nit'].astype(str).str.strip()
    
    def normalizar_documento(doc):
        """Normaliza un documento quitando guiones y convirtiendo a mayúsculas"""
        if pd.isna(doc):
            return ''
        doc_str = str(doc).strip().upper()
        # Quitar guiones y espacios
        doc_str = doc_str.replace('-', '').replace(' ', '')
        return doc_str
    
    # Crear diccionarios de mapeo
    # 1. Diccionario NIT -> NUMERO DE PEDIDO
    pedidos_por_nit = {}
    # 2. Diccionario DOCUMENTO NORMALIZADO -> NUMERO DE PEDIDO
    pedidos_por_doc = {}
    
    for _, row in df_pedidos.iterrows():
        # Convertir NUMERO DE PEDIDO a string sin decimales
        num_pedido = row['NUMERO DE PEDIDO']
        if pd.notna(num_pedido):
            try:
                num_pedido = str(int(float(num_pedido)))
            except:
                num_pedido = str(num_pedido).strip()
        else:
            num_pedido = ''
        
        # Mapeo por NIT
        nit = str(row['IDENTIFICACION']).strip()
        pedidos_por_nit[nit] = num_pedido
        
        # Mapeo por DOCUMENTO ASOCIADO
        if 'DOCUMENTO ASOCIADO' in row and pd.notna(row['DOCUMENTO ASOCIADO']):
            doc_normalizado = normalizar_documento(row['DOCUMENTO ASOCIADO'])
            if doc_normalizado:
                pedidos_por_doc[doc_normalizado] = num_pedido
    
    print(f"Total de NITs en pedidos: {len(pedidos_por_nit)}")
    print(f"Total de DOCUMENTOS en pedidos: {len(pedidos_por_doc)}")
    print(f"Total de registros en planilla inicial: {len(df_inicial)}")
    
    # Actualizar el campo Nrodcto
    registros_actualizados_nit = 0
    registros_actualizados_doc = 0
    registros_no_encontrados = []
    
    for idx, row in df_inicial.iterrows():
        nit = str(row['nit']).strip()
        nrodcto_actual = str(row['Nrodcto'])
        nrodcto_normalizado = normalizar_documento(nrodcto_actual)
        
        num_pedido = None
        metodo = None
        
        # Método 1: Intentar por NIT
        if nit in pedidos_por_nit and pedidos_por_nit[nit]:
            num_pedido = pedidos_por_nit[nit]
            metodo = 'NIT'
            registros_actualizados_nit += 1
        # Método 2: Si no encontró por NIT, intentar por DOCUMENTO
        elif nrodcto_normalizado in pedidos_por_doc and pedidos_por_doc[nrodcto_normalizado]:
            num_pedido = pedidos_por_doc[nrodcto_normalizado]
            metodo = 'DOCUMENTO'
            registros_actualizados_doc += 1
        
        # Si encontró el pedido por algún método, actualizar
        if num_pedido and metodo:
            # Crear el nuevo formato: Nrodcto-NUMERO_DE_PEDIDO
            nuevo_nrodcto = f"{nrodcto_actual}-{num_pedido}"
            df_inicial.at[idx, 'Nrodcto'] = nuevo_nrodcto
        else:
            registros_no_encontrados.append(f"{nit}|{nrodcto_actual}")
    
    total_actualizados = registros_actualizados_nit + registros_actualizados_doc
    print(f"\nRegistros actualizados por NIT: {registros_actualizados_nit}")
    print(f"Registros actualizados por DOCUMENTO: {registros_actualizados_doc}")
    print(f"Total actualizados: {total_actualizados}")
    print(f"Registros sin coincidencia: {len(registros_no_encontrados)}")
    
    if registros_no_encontrados and len(registros_no_encontrados) <= 5:
        print(f"Ejemplos de registros no encontrados: {registros_no_encontrados[:5]}")
    
    return df_inicial


def guardar_con_formato_bogota(df_datos, filas_encabezado):
    """
    Guarda el DataFrame manteniendo el formato original de la planilla inicial.
    Retorna el BytesIO con el Excel generado.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Crear un nuevo workbook
    wb = Workbook()
    ws = wb.active
    
    # Escribir las primeras 4 filas de encabezado originales
    for r_idx, row in enumerate(filas_encabezado.values, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Título
                cell.font = Font(bold=True, size=12)
            elif r_idx == 4:  # Encabezados de columnas
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Escribir los datos actualizados a partir de la fila 5
    for r_idx, row in enumerate(dataframe_to_rows(df_datos, index=False, header=False), start=5):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Guardar en BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


class MailboxHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=os.path.dirname(__file__), **kwargs)
    
    def do_GET(self):
        if self.path == '/' or self.path == '/index.html':
            self.send_html_app()
        else:
            super().do_GET()
    
    def do_POST(self):
        if self.path == '/process':
            self.process_files()
        else:
            self.send_response(404)
            self.end_headers()
    
    def send_html_app(self):
        html_content = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>🚀 Creador de Relaciones Mailbox</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            justify-content: center;
            align-items: center;
            padding: 20px;
        }
        
        .container {
            background: white;
            border-radius: 20px;
            padding: 40px;
            box-shadow: 0 20px 40px rgba(0,0,0,0.1);
            max-width: 600px;
            width: 100%;
        }
        
        .header {
            text-align: center;
            margin-bottom: 30px;
        }
        
        .header h1 {
            color: #2c3e50;
            font-size: 28px;
            font-weight: 700;
            margin-bottom: 10px;
        }
        
        .header p {
            color: #7f8c8d;
            font-size: 16px;
        }
        
        .file-section {
            margin-bottom: 25px;
            padding: 20px;
            border: 2px dashed #e0e0e0;
            border-radius: 10px;
            transition: all 0.3s ease;
        }
        
        .file-section:hover {
            border-color: #3498db;
            background-color: #f8fafb;
        }
        
        .file-section.has-file {
            border-color: #27ae60;
            background-color: #f0fff4;
        }
        
        .file-label {
            display: block;
            font-weight: 600;
            color: #2c3e50;
            margin-bottom: 10px;
            font-size: 16px;
        }
        
        .file-input {
            width: 100%;
            padding: 12px;
            border: 1px solid #ddd;
            border-radius: 8px;
            font-size: 14px;
            background-color: white;
        }
        
        .file-status {
            margin-top: 10px;
            padding: 8px 12px;
            border-radius: 6px;
            font-size: 14px;
            font-weight: 500;
        }
        
        .status-success {
            background-color: #d4edda;
            color: #155724;
            border: 1px solid #c3e6cb;
        }
        
        .status-error {
            background-color: #f8d7da;
            color: #721c24;
            border: 1px solid #f5c6cb;
        }
        
        .process-btn {
            width: 100%;
            padding: 15px;
            background: linear-gradient(135deg, #27ae60, #2ecc71);
            color: white;
            border: none;
            border-radius: 10px;
            font-size: 18px;
            font-weight: 700;
            cursor: pointer;
            transition: all 0.3s ease;
            margin-top: 20px;
        }
        
        .process-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 10px 20px rgba(46, 204, 113, 0.3);
        }
        
        .process-btn:disabled {
            background: #bdc3c7;
            cursor: not-allowed;
            transform: none;
            box-shadow: none;
        }
        
        .result-section {
            margin-top: 30px;
            padding: 20px;
            border-radius: 10px;
            display: none;
        }
        
        .result-success {
            background-color: #d4edda;
            border: 1px solid #c3e6cb;
            color: #155724;
        }
        
        .result-error {
            background-color: #f8d7da;
            border: 1px solid #f5c6cb;
            color: #721c24;
        }
        
        .download-btn {
            display: inline-block;
            margin-top: 15px;
            padding: 10px 20px;
            background-color: #3498db;
            color: white;
            text-decoration: none;
            border-radius: 6px;
            font-weight: 600;
            transition: all 0.3s ease;
        }
        
        .download-btn:hover {
            background-color: #2980b9;
            transform: translateY(-1px);
        }
        
        .small-text {
            font-size: 0.9em;
            color: #666;
            margin-top: 10px;
            font-style: italic;
        }
        
        .loading {
            display: none;
            text-align: center;
            margin-top: 20px;
        }
        
        .spinner {
            border: 4px solid #f3f3f3;
            border-top: 4px solid #3498db;
            border-radius: 50%;
            width: 40px;
            height: 40px;
            animation: spin 1s linear infinite;
            margin: 0 auto 15px;
        }
        
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        
        .info-box {
            background-color: #e8f4fd;
            border-left: 4px solid #3498db;
            padding: 15px;
            margin-bottom: 25px;
            border-radius: 0 8px 8px 0;
        }
        
        .info-box h3 {
            color: #2980b9;
            font-size: 16px;
            margin-bottom: 8px;
        }
        
        .info-box ul {
            color: #34495e;
            font-size: 14px;
            margin-left: 20px;
        }
        
        .info-box li {
            margin-bottom: 5px;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🚀 Creador de Relaciones Mailbox</h1>
            <p>Combina automáticamente planillas CSV de manera fácil y rápida</p>
        </div>
        
        <div style="margin-bottom: 20px;">
            <label class="file-label" for="modoSelector">🔧 Selecciona el modo de operación:</label>
            <select id="modoSelector" class="file-input" style="padding: 12px;">
                <option value="normal">Modo Normal (Madre + Ofimatic)</option>
                <option value="bogota">Modo Bogotá (Relacionar por NIT)</option>
                <option value="filtrar_bogota">Filtrar Bogotá (Solo B-BOGOTA y B-SOACHA)</option>
                <option value="medellin_libro2">Medellín → Libro2 (Formato Ruteo)</option>
                <option value="bogota_libro2">Bogotá → Libro2 (Formato Ruteo)</option>
            </select>
        </div>
        
        <div class="info-box" id="infoNormal">
            <h3>📋 Formato de archivos soportados (Modo Normal):</h3>
            <ul>
                <li><strong>Formatos:</strong> CSV (.csv), Excel (.xlsx, .xls)</li>
                <li><strong>Planilla Madre:</strong> Debe tener columnas identificationPatient e idOrder</li>
                <li><strong>Planilla Ofimatic:</strong> Las primeras 4 filas se omiten automáticamente, debe tener columnas nit y Nrodcto</li>
                <li><strong>CSV:</strong> Se detecta automáticamente el separador (punto y coma o coma)</li>
                <li><strong>Resultado:</strong> Archivo Excel (.xlsx) con filtros automáticos y formato preservado</li>
            </ul>
        </div>
        
        <div class="info-box" id="infoBogota" style="display: none;">
            <h3>📋 Formato de archivos soportados (Modo Bogotá):</h3>
            <ul>
                <li><strong>Planilla Inicial Bogotá:</strong> Archivo Excel con formato base</li>
                <li><strong>Planilla de Pedidos:</strong> Excel con columnas IDENTIFICACION y NUMERO DE PEDIDO</li>
                <li><strong>Proceso:</strong> Relaciona por NIT y actualiza Nrodcto a formato: Nrodcto-NUMERO_PEDIDO</li>
                <li><strong>Resultado:</strong> Excel con el mismo formato de Planilla Inicial Bogotá</li>
            </ul>
        </div>
        
        <div class="info-box" id="infoFiltrarBogota" style="display: none;">
            <h3>📋 Filtrar pedidos de Bogotá (Solo B-BOGOTA y B-SOACHA):</h3>
            <ul>
                <li><strong>Archivo:</strong> PLANILLAS OFMATIC BOGOTA.xlsx (o similar)</li>
                <li><strong>Proceso:</strong> Filtra solo los pedidos con Destino = "B-BOGOTA" o "B-SOACHA"</li>
                <li><strong>Resultado:</strong> Excel filtrado con el mismo formato original</li>
                <li><strong>Nota:</strong> Solo necesitas seleccionar UN archivo</li>
            </ul>
        </div>
        
        <div class="info-box" id="infoMedellinLibro2" style="display: none;">
            <h3>📋 Medellín → Libro2 (Formato Ruteo):</h3>
            <ul>
                <li><strong>Planilla Madre:</strong> Excel con datos de Medellín (identificationPatient, idOrder, addressPatient, phonePatient, cityNameOrder)</li>
                <li><strong>Planilla Ofimatic:</strong> PLANILLAS OFMATIC BOGOTA.xlsx (con nit, Nrodcto, NOMBRE, DIRECCION, TEL1, TEL2, TipoVta, Destino)</li>
                <li><strong>Proceso:</strong> Relaciona por NIT y transforma al formato Libro2.xlsx para ruteo</li>
                <li><strong>Resultado:</strong> Excel con formato: Nombre Vehículo, Título de la Visita, Dirección, ID Referencia, Notas, Teléfono</li>
                <li><strong>Dirección:</strong> Se usa la de la planilla madre + ", " + ciudad + ", Antioquia"</li>
            </ul>
        </div>
        
        <div class="info-box" id="infoBogotaLibro2" style="display: none;">
            <h3>📋 Bogotá → Libro2 (Formato Ruteo):</h3>
            <ul>
                <li><strong>Planilla Ehlpharma:</strong> Excel con datos de Bogotá (IDENTIFICACION, NUMERO DE PEDIDO, DIRECCION DE ENTREGA, CELULAR, CIUDAD DE ENTREGA)</li>
                <li><strong>Planilla Ofimatic:</strong> PLANILLAS OFMATIC BOGOTA.xlsx (con nit, Nrodcto, NomMensajero, NOMBRE, DIRECCION, TEL1, TEL2, TipoVta, Destino)</li>
                <li><strong>Proceso:</strong> Relaciona por NIT y transforma al formato Libro2.xlsx para ruteo en Cundinamarca</li>
                <li><strong>Resultado:</strong> Excel con formato: Nombre Vehículo, Título de la Visita, Dirección, ID Referencia, Notas, Teléfono</li>
                <li><strong>Destino:</strong> Extrae ciudad de formato "B-CIUDAD" → "CIUDAD" en mayúsculas</li>
                <li><strong>Ciudad Entrega:</strong> Extrae de "Zipaquirá-Cundinamarca-Colombia" → "ZIPAQUIRA"</li>
                <li><strong>Dirección:</strong> DIRECCION DE ENTREGA de ehlpharma + ", " + ciudad + ", Cundinamarca"</li>
                <li><strong>Teléfono:</strong> CELULAR de ehlpharma, sino TEL1 o TEL2 de ofimatic</li>
            </ul>
        </div>
        
        <form id="fileForm">
            <div class="file-section" id="madreSection">
                <label class="file-label" for="madreFile" id="madreLabel">1️⃣ Planilla Madre (.csv/.xlsx/.xls)</label>
                <input type="file" id="madreFile" class="file-input" accept=".csv,.xlsx,.xls" required>
                <div id="madreStatus" class="file-status" style="display: none;"></div>
            </div>
            
            <div class="file-section" id="ofimaticSection">
                <label class="file-label" for="ofimaticFile" id="ofimaticLabel">2️⃣ Planilla Ofimatic (.csv/.xlsx/.xls)</label>
                <input type="file" id="ofimaticFile" class="file-input" accept=".csv,.xlsx,.xls" required>
                <div id="ofimaticStatus" class="file-status" style="display: none;"></div>
            </div>
            
            <button type="submit" id="processBtn" class="process-btn">
                3️⃣ ¡GENERAR ARCHIVO COMBINADO!
            </button>
        </form>
        
        <div class="loading" id="loading">
            <div class="spinner"></div>
            <p>🔄 Procesando archivos...</p>
        </div>
        
        <div class="result-section" id="result">
            <div id="resultContent"></div>
        </div>
    </div>

    <script>
        const modoSelector = document.getElementById('modoSelector');
        const madreFile = document.getElementById('madreFile');
        const ofimaticFile = document.getElementById('ofimaticFile');
        const madreLabel = document.getElementById('madreLabel');
        const ofimaticLabel = document.getElementById('ofimaticLabel');
        const processBtn = document.getElementById('processBtn');
        const form = document.getElementById('fileForm');
        const loading = document.getElementById('loading');
        const result = document.getElementById('result');
        const infoNormal = document.getElementById('infoNormal');
        const infoBogota = document.getElementById('infoBogota');
        const infoFiltrarBogota = document.getElementById('infoFiltrarBogota');
        const infoMedellinLibro2 = document.getElementById('infoMedellinLibro2');
        const infoBogotaLibro2 = document.getElementById('infoBogotaLibro2');
        
        // Cambiar etiquetas y descripciones según el modo
        modoSelector.addEventListener('change', () => {
            const modo = modoSelector.value;
            if (modo === 'bogota') {
                madreLabel.textContent = '1️⃣ Planilla Inicial Bogotá (.xlsx)';
                ofimaticLabel.textContent = '2️⃣ Planilla de Pedidos (.xlsx)';
                document.getElementById('madreSection').style.display = 'block';
                document.getElementById('ofimaticSection').style.display = 'block';
                infoNormal.style.display = 'none';
                infoBogota.style.display = 'block';
                infoFiltrarBogota.style.display = 'none';
                infoMedellinLibro2.style.display = 'none';
                infoBogotaLibro2.style.display = 'none';
                processBtn.textContent = '3️⃣ ¡RELACIONAR PLANILLAS BOGOTÁ!';
                madreFile.required = true;
                ofimaticFile.required = true;
            } else if (modo === 'filtrar_bogota') {
                madreLabel.textContent = '1️⃣ Planilla Ofimatic Bogotá (.xlsx)';
                document.getElementById('madreSection').style.display = 'block';
                document.getElementById('ofimaticSection').style.display = 'none';
                infoNormal.style.display = 'none';
                infoBogota.style.display = 'none';
                infoFiltrarBogota.style.display = 'block';
                infoMedellinLibro2.style.display = 'none';
                infoBogotaLibro2.style.display = 'none';
                processBtn.textContent = '2️⃣ ¡FILTRAR PEDIDOS BOGOTÁ!';
                madreFile.required = true;
                ofimaticFile.required = false;
            } else if (modo === 'medellin_libro2') {
                madreLabel.textContent = '1️⃣ Planilla Madre Medellín (.xlsx)';
                ofimaticLabel.textContent = '2️⃣ Planilla Ofimatic (.xlsx)';
                document.getElementById('madreSection').style.display = 'block';
                document.getElementById('ofimaticSection').style.display = 'block';
                infoNormal.style.display = 'none';
                infoBogota.style.display = 'none';
                infoFiltrarBogota.style.display = 'none';
                infoMedellinLibro2.style.display = 'block';
                infoBogotaLibro2.style.display = 'none';
                processBtn.textContent = '3️⃣ ¡GENERAR ARCHIVO LIBRO2!';
                madreFile.required = true;
                ofimaticFile.required = true;
            } else if (modo === 'bogota_libro2') {
                madreLabel.textContent = '1️⃣ Planilla Ehlpharma Bogotá (.xlsx)';
                ofimaticLabel.textContent = '2️⃣ Planilla Ofimatic Bogotá (.xlsx)';
                document.getElementById('madreSection').style.display = 'block';
                document.getElementById('ofimaticSection').style.display = 'block';
                infoNormal.style.display = 'none';
                infoBogota.style.display = 'none';
                infoFiltrarBogota.style.display = 'none';
                infoMedellinLibro2.style.display = 'none';
                infoBogotaLibro2.style.display = 'block';
                processBtn.textContent = '3️⃣ ¡GENERAR ARCHIVO LIBRO2 BOGOTÁ!';
                madreFile.required = true;
                ofimaticFile.required = true;
            } else {
                madreLabel.textContent = '1️⃣ Planilla Madre (.csv/.xlsx/.xls)';
                ofimaticLabel.textContent = '2️⃣ Planilla Ofimatic (.csv/.xlsx/.xls)';
                document.getElementById('madreSection').style.display = 'block';
                document.getElementById('ofimaticSection').style.display = 'block';
                infoNormal.style.display = 'block';
                infoBogota.style.display = 'none';
                infoFiltrarBogota.style.display = 'none';
                infoMedellinLibro2.style.display = 'none';
                infoBogotaLibro2.style.display = 'none';
                processBtn.textContent = '3️⃣ ¡GENERAR ARCHIVO COMBINADO!';
                madreFile.required = true;
                ofimaticFile.required = true;
            }
            // Reset archivos
            madreFile.value = '';
            ofimaticFile.value = '';
            document.getElementById('madreStatus').style.display = 'none';
            document.getElementById('ofimaticStatus').style.display = 'none';
            document.getElementById('madreSection').classList.remove('has-file');
            document.getElementById('ofimaticSection').classList.remove('has-file');
            checkFormReady();
        });
        
        // Escuchar cambios en el modo también
        modoSelector.addEventListener('change', checkFormReady);
        
        function updateFileStatus(fileInput, statusDiv, sectionDiv) {
            const file = fileInput.files[0];
            if (file) {
                statusDiv.textContent = `✅ Archivo seleccionado: ${file.name}`;
                statusDiv.className = 'file-status status-success';
                statusDiv.style.display = 'block';
                sectionDiv.classList.add('has-file');
            } else {
                statusDiv.style.display = 'none';
                sectionDiv.classList.remove('has-file');
            }
            checkFormReady();
        }
        
        function checkFormReady() {
            const modo = modoSelector.value;
            const madreReady = madreFile.files.length > 0;
            const ofimaticReady = ofimaticFile.files.length > 0;
            
            if (modo === 'filtrar_bogota') {
                processBtn.disabled = !madreReady;
            } else {
                processBtn.disabled = !(madreReady && ofimaticReady);
            }
        }
        
        madreFile.addEventListener('change', () => {
            updateFileStatus(madreFile, document.getElementById('madreStatus'), document.getElementById('madreSection'));
        });
        
        ofimaticFile.addEventListener('change', () => {
            updateFileStatus(ofimaticFile, document.getElementById('ofimaticStatus'), document.getElementById('ofimaticSection'));
        });
        
        form.addEventListener('submit', async (e) => {
            e.preventDefault();
            
            const modo = modoSelector.value;
            
            // Validar según el modo
            if (modo === 'filtrar_bogota') {
                if (!madreFile.files[0]) {
                    alert('Por favor, selecciona el archivo de planilla Ofimatic Bogotá.');
                    return;
                }
            } else {
                if (!madreFile.files[0] || !ofimaticFile.files[0]) {
                    alert('Por favor, selecciona ambos archivos antes de continuar.');
                    return;
                }
            }
            
            // Mostrar loading
            form.style.display = 'none';
            loading.style.display = 'block';
            result.style.display = 'none';
            
            try {
                const formData = new FormData();
                formData.append('madre', madreFile.files[0]);
                if (modo !== 'filtrar_bogota' && ofimaticFile.files[0]) {
                    formData.append('ofimatic', ofimaticFile.files[0]);
                }
                formData.append('modo', modo);
                
                const response = await fetch('/process', {
                    method: 'POST',
                    body: formData
                });
                
                const data = await response.json();
                
                loading.style.display = 'none';
                result.style.display = 'block';
                
                if (data.success) {
                    result.className = 'result-section result-success';
                    
                    // Crear un blob para el archivo Excel
                    const binaryString = atob(data.excel_data);
                    const bytes = new Uint8Array(binaryString.length);
                    for (let i = 0; i < binaryString.length; i++) {
                        bytes[i] = binaryString.charCodeAt(i);
                    }
                    const blob = new Blob([bytes], { type: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' });
                    const url = URL.createObjectURL(blob);
                    
                    document.getElementById('resultContent').innerHTML = `
                        <h3>🎉 ¡Proceso completado exitosamente!</h3>
                        <p>${data.message}</p>
                        <a href="${url}" 
                           download="${data.filename}" 
                           class="download-btn">
                           📥 Descargar Archivo Excel con Filtros
                        </a>
                        <p class="small-text">El archivo Excel conserva el formato exacto del original de Ofimatic con filtros automáticos</p>
                    `;
                } else {
                    result.className = 'result-section result-error';
                    document.getElementById('resultContent').innerHTML = `
                        <h3>❌ Error en el procesamiento</h3>
                        <p><strong>Error:</strong> ${data.error}</p>
                        <p><strong>Detalles:</strong> ${data.details || 'No hay detalles adicionales.'}</p>
                    `;
                }
            } catch (error) {
                loading.style.display = 'none';
                result.style.display = 'block';
                result.className = 'result-section result-error';
                document.getElementById('resultContent').innerHTML = `
                    <h3>❌ Error de conexión</h3>
                    <p>No se pudo procesar la solicitud: ${error.message}</p>
                `;
            }
            
            form.style.display = 'block';
        });
        
        // Inicializar estado
        checkFormReady();
    </script>
</body>
</html>
        """
        
        self.send_response(200)
        self.send_header('Content-type', 'text/html; charset=utf-8')
        self.send_header('Content-length', len(html_content.encode('utf-8')))
        self.end_headers()
        self.wfile.write(html_content.encode('utf-8'))
    
    def process_files(self):
        try:
            # Obtener el tipo de contenido
            content_type = self.headers['content-type']
            if not content_type.startswith('multipart/form-data'):
                self.send_json_response({'success': False, 'error': 'Tipo de contenido no válido'})
                return
            
            # Obtener la longitud del contenido
            content_length = int(self.headers['content-length'])
            post_data = self.rfile.read(content_length)
            
            # Parsear multipart data (simplificado)
            boundary = content_type.split('boundary=')[1].encode()
            parts = post_data.split(b'--' + boundary)
            
            files = {}
            filenames = {}
            modo = 'normal'  # modo por defecto
            
            for part in parts:
                if b'Content-Disposition' in part:
                    header_end = part.find(b'\r\n\r\n')
                    if header_end != -1:
                        header = part[:header_end].decode('utf-8')
                        content = part[header_end + 4:]
                        
                        # Remover trailing boundary
                        if content.endswith(b'\r\n'):
                            content = content[:-2]
                        
                        # Extraer el nombre del archivo si existe
                        if b'filename=' in part:
                            filename = ''
                            if 'filename="' in header:
                                start = header.find('filename="') + 10
                                end = header.find('"', start)
                                filename = header[start:end]
                            
                            if 'name="madre"' in header:
                                files['madre'] = content
                                filenames['madre'] = filename
                            elif 'name="ofimatic"' in header:
                                files['ofimatic'] = content
                                filenames['ofimatic'] = filename
                        elif 'name="modo"' in header:
                            # Extraer el valor del modo
                            modo = content.decode('utf-8').strip()
            
            if 'madre' not in files:
                self.send_json_response({
                    'success': False, 
                    'error': 'No se pudo leer el archivo',
                    'details': 'Asegúrate de que el archivo está seleccionado'
                })
                return
            
            # Procesar los archivos según el modo
            if modo == 'filtrar_bogota':
                result = self.process_filtrar_bogota(
                    files['madre'], filenames['madre']
                )
            elif modo == 'bogota':
                if 'ofimatic' not in files:
                    self.send_json_response({
                        'success': False, 
                        'error': 'No se pudieron leer los archivos',
                        'details': 'Asegúrate de que ambos archivos están seleccionados'
                    })
                    return
                result = self.process_bogota_files(
                    files['madre'], filenames['madre'],
                    files['ofimatic'], filenames['ofimatic']
                )
            elif modo == 'medellin_libro2':
                if 'ofimatic' not in files:
                    self.send_json_response({
                        'success': False, 
                        'error': 'No se pudieron leer los archivos',
                        'details': 'Asegúrate de que ambos archivos están seleccionados'
                    })
                    return
                result = self.process_medellin_libro2(
                    files['madre'], filenames['madre'],
                    files['ofimatic'], filenames['ofimatic']
                )
            elif modo == 'bogota_libro2':
                if 'ofimatic' not in files:
                    self.send_json_response({
                        'success': False, 
                        'error': 'No se pudieron leer los archivos',
                        'details': 'Asegúrate de que ambos archivos están seleccionados'
                    })
                    return
                result = self.process_bogota_libro2(
                    files['madre'], filenames['madre'],
                    files['ofimatic'], filenames['ofimatic']
                )
            else:
                if 'ofimatic' not in files:
                    self.send_json_response({
                        'success': False, 
                        'error': 'No se pudieron leer los archivos',
                        'details': 'Asegúrate de que ambos archivos están seleccionados'
                    })
                    return
                result = self.process_data_files(
                    files['madre'], filenames['madre'],
                    files['ofimatic'], filenames['ofimatic']
                )
            self.send_json_response(result)
            
        except Exception as e:
            self.send_json_response({
                'success': False, 
                'error': f'Error inesperado: {str(e)}',
                'details': 'Verifica que los archivos tengan el formato correcto'
            })
    
    def process_data_files(self, madre_content, madre_filename, ofimatic_content, ofimatic_filename):
        try:
            print(f"🔄 Procesando archivos: {madre_filename} y {ofimatic_filename}")
            
            # Leer planilla madre
            if madre_filename.lower().endswith(('.xlsx', '.xls')):
                df_madre = leer_excel_inteligente_desde_contenido(madre_content)
            else:
                # Para CSV, intentar decodificar con diferentes codificaciones
                codificaciones = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
                df_madre = None
                
                for encoding in codificaciones:
                    try:
                        madre_text = madre_content.decode(encoding)
                        try:
                            df_madre = pd.read_csv(StringIO(madre_text), delimiter=';')
                            print(f"✅ Archivo madre leído con codificación: {encoding} y delimitador ';'")
                            break
                        except:
                            try:
                                df_madre = pd.read_csv(StringIO(madre_text), delimiter=',')
                                print(f"✅ Archivo madre leído con codificación: {encoding} y delimitador ','")
                                break
                            except:
                                try:
                                    df_madre = pd.read_csv(StringIO(madre_text))
                                    print(f"✅ Archivo madre leído con codificación: {encoding} y delimitador automático")
                                    break
                                except:
                                    continue
                    except:
                        continue
                
                if df_madre is None:
                    return {
                        'success': False,
                        'error': 'No se pudo leer el archivo madre',
                        'details': 'Verifica que el archivo tenga una codificación válida (UTF-8, Latin-1, etc.)'
                    }
            
            print(f"✅ Planilla madre leída: {len(df_madre)} filas")
            
            # Leer planilla ofimatic usando la función específica
            try:
                df_ofimatic = leer_archivo_ofimatic_desde_contenido(ofimatic_content, ofimatic_filename)
            except Exception as e:
                return {
                    'success': False,
                    'error': f'Error al leer archivo ofimatic: {str(e)}',
                    'details': 'Verifica que el archivo tenga las columnas nit y Nrodcto'
                }
            
            print(f"✅ Planilla ofimatic leída: {len(df_ofimatic)} filas")
            
            # Verificar columnas requeridas
            required_madre_cols = ['identificationPatient', 'idOrder']
            missing_madre = [col for col in required_madre_cols if col not in df_madre.columns]
            if missing_madre:
                return {
                    'success': False,
                    'error': f'Columnas faltantes en planilla madre: {missing_madre}',
                    'details': f'Columnas disponibles: {list(df_madre.columns)}'
                }
            
            required_ofimatic_cols = ['nit', 'Nrodcto']
            missing_ofimatic = [col for col in required_ofimatic_cols if col not in df_ofimatic.columns]
            if missing_ofimatic:
                return {
                    'success': False,
                    'error': f'Columnas faltantes en planilla ofimatic: {missing_ofimatic}',
                    'details': f'Columnas disponibles: {list(df_ofimatic.columns)}'
                }
            
            # Seleccionar y limpiar columnas
            df_madre_reducido = df_madre[['identificationPatient', 'idOrder']].copy()
            df_madre_reducido['identificationPatient'] = df_madre_reducido['identificationPatient'].astype(str)
            df_ofimatic['nit'] = df_ofimatic['nit'].astype(str)
            
            print(f"🔗 Procesando datos...")
            
            # En lugar de crear un DataFrame fusionado, editamos directamente el original
            # Crear un diccionario de mapeo nit -> idOrder
            mapeo_nit_idorder = df_madre_reducido.set_index('identificationPatient')['idOrder'].to_dict()
            
            # Editar directamente el DataFrame de ofimatic
            df_ofimatic['idOrder_mapeado'] = df_ofimatic['nit'].map(mapeo_nit_idorder).fillna('')
            df_ofimatic['idOrder_mapeado'] = df_ofimatic['idOrder_mapeado'].astype(str)
            
            # Limpiar idOrder_mapeado para evitar decimales
            df_ofimatic['idOrder_mapeado'] = df_ofimatic['idOrder_mapeado'].apply(
                lambda x: str(int(float(x))) if x and x.replace('.','',1).isdigit() else x
            )
            
            # Actualizar la columna Nrodcto DIRECTAMENTE en el DataFrame original
            df_ofimatic['Nrodcto'] = df_ofimatic['Nrodcto'].astype(str) + '-' + df_ofimatic['idOrder_mapeado']
            
            # Eliminar la columna temporal
            df_ofimatic = df_ofimatic.drop(columns=['idOrder_mapeado'])
            
            # Generar archivo Excel preservando el formato original
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df_ofimatic.to_excel(writer, sheet_name='Sheet1', index=False)
                
                # Obtener el workbook y worksheet para agregar filtros
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']
                
                # Aplicar filtros automáticos (igual que el original)
                worksheet.auto_filter.ref = f"A1:{chr(65 + len(df_ofimatic.columns) - 1)}{len(df_ofimatic) + 1}"
                
                # Ajustar el ancho de las columnas automáticamente
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            excel_buffer.seek(0)
            excel_data = base64.b64encode(excel_buffer.read()).decode('utf-8')
            
            print(f"✅ Proceso completado: {len(df_ofimatic)} filas en el resultado")
            
            return {
                'success': True,
                'message': f'Archivo procesado exitosamente. {len(df_ofimatic)} registros actualizados.',
                'excel_data': excel_data,
                'filename': 'relaciones_unidas.xlsx'
            }
            
        except Exception as e:
            print(f"❌ Error: {e}")
            return {
                'success': False,
                'error': f'Error al procesar: {str(e)}',
                'details': 'Verifica que los archivos tengan el formato correcto (CSV, XLS, XLSX)'
            }
    
    def process_bogota_files(self, inicial_content, inicial_filename, pedidos_content, pedidos_filename):
        """
        Procesa archivos para el modo Bogotá: relaciona por NIT y actualiza Nrodcto
        """
        try:
            print(f"🔄 [BOGOTÁ] Procesando archivos: {inicial_filename} y {pedidos_filename}")
            
            # Leer planilla inicial de Bogotá
            df_inicial, filas_encabezado, encabezados = leer_planilla_inicial_bogota(inicial_content)
            print(f"✅ Planilla inicial leída: {len(df_inicial)} filas")
            
            # Leer planilla de pedidos
            df_pedidos = leer_planilla_pedidos_bogota(pedidos_content)
            print(f"✅ Planilla de pedidos leída: {len(df_pedidos)} filas")
            
            # Verificar columnas requeridas
            if 'nit' not in df_inicial.columns or 'Nrodcto' not in df_inicial.columns:
                return {
                    'success': False,
                    'error': 'Planilla Inicial no tiene las columnas necesarias (nit, Nrodcto)',
                    'details': f'Columnas disponibles: {list(df_inicial.columns)}'
                }
            
            if 'IDENTIFICACION' not in df_pedidos.columns or 'NUMERO DE PEDIDO' not in df_pedidos.columns:
                return {
                    'success': False,
                    'error': 'Planilla de Pedidos no tiene las columnas necesarias (IDENTIFICACION, NUMERO DE PEDIDO)',
                    'details': f'Columnas disponibles: {list(df_pedidos.columns)}'
                }
            
            # Relacionar por NIT
            print("🔗 Relacionando datos por NIT...")
            df_actualizado = relacionar_por_nit_bogota(df_inicial, df_pedidos)
            
            # Guardar con formato original
            print("💾 Generando archivo Excel con formato original...")
            excel_buffer = guardar_con_formato_bogota(df_actualizado, filas_encabezado)
            
            excel_data = base64.b64encode(excel_buffer.read()).decode('utf-8')
            
            print(f"✅ Proceso completado: {len(df_actualizado)} registros en el resultado")
            
            from datetime import datetime
            fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            return {
                'success': True,
                'message': f'Archivo procesado exitosamente. {len(df_actualizado)} registros procesados.',
                'excel_data': excel_data,
                'filename': f'Planilla_Relacionada_Bogota_{fecha_actual}.xlsx'
            }
            
        except Exception as e:
            print(f"❌ Error: {e}")
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'error': f'Error al procesar archivos de Bogotá: {str(e)}',
                'details': 'Verifica que los archivos sean Excel (.xlsx) y tengan las columnas correctas'
            }
    
    def process_filtrar_bogota(self, archivo_content, archivo_filename):
        """
        Filtra la planilla Ofimatic Bogotá para dejar solo los pedidos con Destino = B-BOGOTA
        """
        try:
            print(f"🔄 [FILTRAR BOGOTÁ] Procesando archivo: {archivo_filename}")
            
            # Leer el archivo completo sin procesar
            df_completo = pd.read_excel(BytesIO(archivo_content), header=None)
            
            # Los encabezados están en la fila 3 (índice 3)
            encabezados = df_completo.iloc[3].tolist()
            
            # Guardar las primeras 4 filas para mantener el formato original
            filas_encabezado = df_completo.iloc[0:4]
            
            # Los datos comienzan desde la fila 4 (índice 4)
            df_datos = pd.read_excel(BytesIO(archivo_content), skiprows=4)
            df_datos.columns = encabezados
            
            print(f"✅ Planilla leída: {len(df_datos)} registros totales")
            
            # Filtrar solo los registros con Destino = B-BOGOTA o B-SOACHA
            if 'Destino' not in df_datos.columns:
                return {
                    'success': False,
                    'error': 'El archivo no tiene la columna "Destino"',
                    'details': f'Columnas disponibles: {list(df_datos.columns)}'
                }
            
            df_filtrado = df_datos[df_datos['Destino'].isin(['B-BOGOTA', 'B-SOACHA'])].copy()
            
            # Contar registros por destino
            count_bogota = len(df_filtrado[df_filtrado['Destino'] == 'B-BOGOTA'])
            count_soacha = len(df_filtrado[df_filtrado['Destino'] == 'B-SOACHA'])
            print(f"✅ Registros filtrados: {len(df_filtrado)} ({count_bogota} B-BOGOTA, {count_soacha} B-SOACHA)")
            
            if len(df_filtrado) == 0:
                return {
                    'success': False,
                    'error': 'No se encontraron registros con Destino = B-BOGOTA o B-SOACHA',
                    'details': f'Total de registros en el archivo: {len(df_datos)}'
                }
            
            # Guardar con formato original
            print("💾 Generando archivo Excel con formato original...")
            excel_buffer = guardar_con_formato_bogota(df_filtrado, filas_encabezado)
            
            excel_data = base64.b64encode(excel_buffer.read()).decode('utf-8')
            
            print(f"✅ Proceso completado: {len(df_filtrado)} registros en el resultado")
            
            from datetime import datetime
            fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            return {
                'success': True,
                'message': f'Archivo filtrado exitosamente. {len(df_filtrado)} registros ({count_bogota} B-BOGOTA, {count_soacha} B-SOACHA) de {len(df_datos)} totales.',
                'excel_data': excel_data,
                'filename': f'Planilla_Filtrada_BOGOTA_SOACHA_{fecha_actual}.xlsx'
            }
            
        except Exception as e:
            print(f"❌ Error: {e}")
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'error': f'Error al filtrar archivo: {str(e)}',
                'details': 'Verifica que el archivo sea Excel (.xlsx) y tenga la estructura correcta'
            }
    
    def process_medellin_libro2(self, madre_content, madre_filename, ofimatic_content, ofimatic_filename):
        """
        Procesa archivos de Medellín (madre + ofimatic) y los transforma al formato Libro2.xlsx
        """
        try:
            print(f"🔄 [MEDELLÍN → LIBRO2] Procesando archivos: {madre_filename} y {ofimatic_filename}")
            
            # Leer planilla madre
            df_madre = leer_excel_inteligente_desde_contenido(madre_content)
            print(f"✅ Planilla madre leída: {len(df_madre)} filas")
            print(f"   Columnas disponibles: {list(df_madre.columns)}")
            
            # Leer planilla ofimatic (con estructura especial de 4 filas de encabezado)
            df_ofimatic = pd.read_excel(BytesIO(ofimatic_content), header=3)
            print(f"✅ Planilla ofimatic leída: {len(df_ofimatic)} filas")
            print(f"   Columnas disponibles: {list(df_ofimatic.columns)}")
            
            # Verificar columnas requeridas en planilla madre
            required_madre_cols = ['identificationPatient', 'idOrder']
            optional_madre_cols = ['addressPatient', 'mobilePhonePatient', 'cityNameOrder']
            missing_madre = [col for col in required_madre_cols if col not in df_madre.columns]
            if missing_madre:
                return {
                    'success': False,
                    'error': f'Columnas faltantes en planilla madre: {missing_madre}',
                    'details': f'Columnas disponibles: {list(df_madre.columns)}'
                }
            
            # Verificar columnas requeridas en planilla ofimatic
            required_ofimatic_cols = ['nit', 'Nrodcto']
            optional_ofimatic_cols = ['NomMensajero', 'NOMBRE', 'DIRECCION', 'TEL1', 'TEL2', 'TipoVta', 'Destino']
            missing_ofimatic = [col for col in required_ofimatic_cols if col not in df_ofimatic.columns]
            if missing_ofimatic:
                return {
                    'success': False,
                    'error': f'Columnas faltantes en planilla ofimatic: {missing_ofimatic}',
                    'details': f'Columnas disponibles: {list(df_ofimatic.columns)}'
                }
            
            # Normalizar tipos de datos
            df_madre['identificationPatient'] = df_madre['identificationPatient'].astype(str)
            df_ofimatic['nit'] = df_ofimatic['nit'].astype(str)
            
            # Paso 1: Relacionar por NIT (igual que en Medellín normal)
            print("🔗 Paso 1: Relacionando por NIT...")
            mapeo_nit_idorder = df_madre.set_index('identificationPatient')['idOrder'].to_dict()
            
            # Crear diccionarios de mapeo adicionales desde la planilla madre
            mapeo_address = {}
            mapeo_phone = {}
            mapeo_city = {}
            
            if 'addressPatient' in df_madre.columns:
                mapeo_address = df_madre.set_index('identificationPatient')['addressPatient'].to_dict()
            if 'mobilePhonePatient' in df_madre.columns:
                mapeo_phone = df_madre.set_index('identificationPatient')['mobilePhonePatient'].to_dict()
            if 'cityNameOrder' in df_madre.columns:
                mapeo_city = df_madre.set_index('identificationPatient')['cityNameOrder'].to_dict()
            
            # Aplicar mapeos
            df_ofimatic['idOrder_mapeado'] = df_ofimatic['nit'].map(mapeo_nit_idorder).fillna('')
            df_ofimatic['addressPatient_madre'] = df_ofimatic['nit'].map(mapeo_address).fillna('')
            df_ofimatic['phonePatient_madre'] = df_ofimatic['nit'].map(mapeo_phone).fillna('')
            df_ofimatic['cityNameOrder_madre'] = df_ofimatic['nit'].map(mapeo_city).fillna('')
            
            # Limpiar idOrder_mapeado para evitar decimales
            df_ofimatic['idOrder_mapeado'] = df_ofimatic['idOrder_mapeado'].apply(
                lambda x: str(int(float(x))) if x and str(x).replace('.','',1).replace('-','',1).isdigit() else str(x)
            )
            
            # Actualizar Nrodcto con el formato: Nrodcto-idOrder
            df_ofimatic['Nrodcto_relacionado'] = df_ofimatic.apply(
                lambda row: f"{row['Nrodcto']}-{row['idOrder_mapeado']}" if row['idOrder_mapeado'] else row['Nrodcto'],
                axis=1
            )
            
            print(f"✅ Relacionados: {(df_ofimatic['idOrder_mapeado'] != '').sum()} de {len(df_ofimatic)} registros")
            
            # Paso 2: Transformar al formato Libro2.xlsx
            print("🔄 Paso 2: Transformando al formato Libro2...")
            
            # Crear DataFrame con estructura de Libro2
            df_libro2 = pd.DataFrame()
            
            # Nombre Vehiculo = NomMensajero (de ofimatic)
            df_libro2['Nombre Vehiculo'] = df_ofimatic['NomMensajero'] if 'NomMensajero' in df_ofimatic.columns else ''
            
            # Título de la Visita = NOMBRE (de ofimatic)
            df_libro2['Título de la Visita'] = df_ofimatic['NOMBRE'] if 'NOMBRE' in df_ofimatic.columns else ''
            
            # Dirección = addressPatient de madre (si existe) + ", " + ciudad + ", Antioquia"
            # Si no hay dirección de madre, usar DIRECCION de ofimatic
            df_libro2['Dirección'] = df_ofimatic.apply(
                lambda row: self._construir_direccion(row, mapeo_address, mapeo_city),
                axis=1
            )
            
            # Latitud y Longitud - vacíos
            df_libro2['Latitud'] = None
            df_libro2['Longitud'] = None
            
            # ID Referencia = Nrodcto relacionado
            df_libro2['ID Referencia'] = df_ofimatic['Nrodcto_relacionado']
            
            # Notas = TipoVta (de ofimatic)
            df_libro2['Notas'] = df_ofimatic['TipoVta'] if 'TipoVta' in df_ofimatic.columns else ''
            
            # Persona de Contacto - vacío
            df_libro2['Persona de Contacto'] = None
            
            # Teléfono = mobilePhonePatient de madre si existe, sino TEL1 o TEL2 de ofimatic
            df_libro2['Teléfono'] = df_ofimatic.apply(
                lambda row: self._obtener_telefono_medellin(row),
                axis=1
            )
            
            # Emails - vacío
            df_libro2['Emails'] = None
            
            print(f"✅ DataFrame Libro2 creado: {len(df_libro2)} registros")
            
            # Generar archivo Excel
            print("💾 Generando archivo Excel formato Libro2...")
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df_libro2.to_excel(writer, sheet_name='Hoja1', index=False)
                
                # Ajustar anchos de columna
                workbook = writer.book
                worksheet = writer.sheets['Hoja1']
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            excel_buffer.seek(0)
            excel_data = base64.b64encode(excel_buffer.read()).decode('utf-8')
            
            print(f"✅ Proceso completado: {len(df_libro2)} registros en formato Libro2")
            
            from datetime import datetime
            fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            return {
                'success': True,
                'message': f'Archivo transformado exitosamente. {len(df_libro2)} registros en formato Libro2.',
                'excel_data': excel_data,
                'filename': f'Libro2_Medellin_{fecha_actual}.xlsx'
            }
            
        except Exception as e:
            print(f"❌ Error: {e}")
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'error': f'Error al transformar a Libro2: {str(e)}',
                'details': 'Verifica que los archivos tengan las columnas correctas'
            }
    
    def process_bogota_libro2(self, ehlpharma_content, ehlpharma_filename, ofimatic_content, ofimatic_filename):
        """
        Procesa archivos de Bogotá (ehlpharma + ofimatic) y los transforma al formato Libro2.xlsx
        """
        try:
            print(f"🔄 [BOGOTÁ → LIBRO2] Procesando archivos: {ehlpharma_filename} y {ofimatic_filename}")
            
            # Leer planilla ehlpharma (similar a madre pero con columnas diferentes)
            df_ehlpharma = leer_excel_inteligente_desde_contenido(ehlpharma_content)
            print(f"✅ Planilla ehlpharma leída: {len(df_ehlpharma)} filas")
            print(f"   Columnas disponibles: {list(df_ehlpharma.columns)}")
            
            # Leer planilla ofimatic (con estructura especial de 4 filas de encabezado)
            df_ofimatic = pd.read_excel(BytesIO(ofimatic_content), header=3)
            print(f"✅ Planilla ofimatic leída: {len(df_ofimatic)} filas")
            print(f"   Columnas disponibles: {list(df_ofimatic.columns)}")
            
            # Verificar columnas requeridas en planilla ehlpharma
            required_ehlpharma_cols = ['IDENTIFICACION', 'NUMERO DE PEDIDO', 'DOCUMENTO ASOCIADO']
            optional_ehlpharma_cols = ['DIRECCION DE ENTREGA', 'CELULAR', 'CIUDAD DE ENTREGA']
            missing_ehlpharma = [col for col in required_ehlpharma_cols if col not in df_ehlpharma.columns]
            if missing_ehlpharma:
                return {
                    'success': False,
                    'error': f'Columnas faltantes en planilla ehlpharma: {missing_ehlpharma}',
                    'details': f'Columnas disponibles: {list(df_ehlpharma.columns)}'
                }
            
            # Verificar columnas requeridas en planilla ofimatic
            required_ofimatic_cols = ['nit', 'Nrodcto']
            optional_ofimatic_cols = ['NomMensajero', 'NOMBRE', 'DIRECCION', 'TEL1', 'TEL2', 'TipoVta', 'Destino']
            missing_ofimatic = [col for col in required_ofimatic_cols if col not in df_ofimatic.columns]
            if missing_ofimatic:
                return {
                    'success': False,
                    'error': f'Columnas faltantes en planilla ofimatic: {missing_ofimatic}',
                    'details': f'Columnas disponibles: {list(df_ofimatic.columns)}'
                }
            
            # Normalizar tipos de datos
            df_ehlpharma['IDENTIFICACION'] = df_ehlpharma['IDENTIFICACION'].astype(str)
            df_ehlpharma['DOCUMENTO ASOCIADO'] = df_ehlpharma['DOCUMENTO ASOCIADO'].apply(
                lambda x: self._normalizar_documento_asociado(x)
            )
            df_ofimatic['nit'] = df_ofimatic['nit'].astype(str)
            
            # Normalizar Nrodcto en ofimatic para facilitar relación
            df_ofimatic['Nrodcto_normalizado'] = df_ofimatic['Nrodcto'].apply(
                lambda x: self._normalizar_documento_asociado(x)
            )
            
            # Paso 1: Relacionar por NIT Y por DOCUMENTO ASOCIADO
            print("🔗 Paso 1: Relacionando por NIT y DOCUMENTO ASOCIADO...")
            
            # Mapeo principal por NIT
            mapeo_nit_idorder = df_ehlpharma.set_index('IDENTIFICACION')['NUMERO DE PEDIDO'].to_dict()
            mapeo_nit_documento = df_ehlpharma.set_index('IDENTIFICACION')['DOCUMENTO ASOCIADO'].to_dict()
            
            # Mapeo secundario por DOCUMENTO ASOCIADO (para relacionar con Nrodcto)
            mapeo_documento_idorder = df_ehlpharma.set_index('DOCUMENTO ASOCIADO')['NUMERO DE PEDIDO'].to_dict()
            # Para obtener el documento asociado original, no necesitamos este mapeo ya que ya está normalizado
            
            # Crear diccionarios de mapeo desde ehlpharma
            # Mapeos por NIT (IDENTIFICACION)
            mapeo_address_nit = {}
            mapeo_phone_nit = {}
            mapeo_city_nit = {}
            
            # Mapeos por DOCUMENTO ASOCIADO
            mapeo_address_doc = {}
            mapeo_phone_doc = {}
            mapeo_city_doc = {}
            
            if 'DIRECCION DE ENTREGA' in df_ehlpharma.columns:
                mapeo_address_nit = df_ehlpharma.set_index('IDENTIFICACION')['DIRECCION DE ENTREGA'].to_dict()
                mapeo_address_doc = df_ehlpharma.set_index('DOCUMENTO ASOCIADO')['DIRECCION DE ENTREGA'].to_dict()
            if 'CELULAR' in df_ehlpharma.columns:
                mapeo_phone_nit = df_ehlpharma.set_index('IDENTIFICACION')['CELULAR'].to_dict()
                mapeo_phone_doc = df_ehlpharma.set_index('DOCUMENTO ASOCIADO')['CELULAR'].to_dict()
            if 'CIUDAD DE ENTREGA' in df_ehlpharma.columns:
                mapeo_city_nit = df_ehlpharma.set_index('IDENTIFICACION')['CIUDAD DE ENTREGA'].to_dict()
                mapeo_city_doc = df_ehlpharma.set_index('DOCUMENTO ASOCIADO')['CIUDAD DE ENTREGA'].to_dict()
            
            # Aplicar mapeos - Intentar primero por NIT, luego por DOCUMENTO ASOCIADO
            def obtener_idorder(row):
                nit = str(row['nit'])
                nrodcto_norm = row['Nrodcto_normalizado']
                
                # Prioridad 1: Por NIT
                if nit in mapeo_nit_idorder:
                    return mapeo_nit_idorder[nit]
                # Prioridad 2: Por DOCUMENTO ASOCIADO (Nrodcto normalizado)
                elif nrodcto_norm in mapeo_documento_idorder:
                    return mapeo_documento_idorder[nrodcto_norm]
                else:
                    return ''
            
            def obtener_documento_asociado(row):
                nit = str(row['nit'])
                nrodcto_norm = row['Nrodcto_normalizado']
                
                # Prioridad 1: Por NIT
                if nit in mapeo_nit_documento:
                    return mapeo_nit_documento[nit]
                # Prioridad 2: Por DOCUMENTO ASOCIADO normalizado (usar el mismo valor normalizado)
                # Si el Nrodcto normalizado coincide con algún documento asociado normalizado
                elif nrodcto_norm in mapeo_documento_idorder:
                    # El documento asociado ya está normalizado en el DataFrame
                    return nrodcto_norm
                else:
                    # Si no hay relación, usar el Nrodcto normalizado
                    return nrodcto_norm
            
            def obtener_address_ehlpharma(row):
                nit = str(row['nit'])
                nrodcto_norm = row['Nrodcto_normalizado']
                
                # Prioridad 1: Por NIT
                if nit in mapeo_address_nit and mapeo_address_nit[nit]:
                    return mapeo_address_nit[nit]
                # Prioridad 2: Por DOCUMENTO ASOCIADO
                elif nrodcto_norm in mapeo_address_doc and mapeo_address_doc[nrodcto_norm]:
                    return mapeo_address_doc[nrodcto_norm]
                else:
                    return ''
            
            def obtener_phone_ehlpharma(row):
                nit = str(row['nit'])
                nrodcto_norm = row['Nrodcto_normalizado']
                
                # Prioridad 1: Por NIT
                if nit in mapeo_phone_nit and mapeo_phone_nit[nit]:
                    return mapeo_phone_nit[nit]
                # Prioridad 2: Por DOCUMENTO ASOCIADO
                elif nrodcto_norm in mapeo_phone_doc and mapeo_phone_doc[nrodcto_norm]:
                    return mapeo_phone_doc[nrodcto_norm]
                else:
                    return ''
            
            def obtener_city_ehlpharma(row):
                nit = str(row['nit'])
                nrodcto_norm = row['Nrodcto_normalizado']
                
                # Prioridad 1: Por NIT
                if nit in mapeo_city_nit and mapeo_city_nit[nit]:
                    return mapeo_city_nit[nit]
                # Prioridad 2: Por DOCUMENTO ASOCIADO
                elif nrodcto_norm in mapeo_city_doc and mapeo_city_doc[nrodcto_norm]:
                    return mapeo_city_doc[nrodcto_norm]
                else:
                    return ''
            
            df_ofimatic['idOrder_mapeado'] = df_ofimatic.apply(obtener_idorder, axis=1)
            df_ofimatic['documento_asociado_mapeado'] = df_ofimatic.apply(obtener_documento_asociado, axis=1)
            df_ofimatic['address_ehlpharma'] = df_ofimatic.apply(obtener_address_ehlpharma, axis=1)
            df_ofimatic['phone_ehlpharma'] = df_ofimatic.apply(obtener_phone_ehlpharma, axis=1)
            df_ofimatic['city_ehlpharma'] = df_ofimatic.apply(obtener_city_ehlpharma, axis=1)
            
            # Limpiar idOrder_mapeado para evitar decimales y NaN
            def limpiar_idorder(x):
                if not x or pd.isna(x) or str(x).strip() == '' or str(x).lower() == 'nan':
                    return ''
                try:
                    if str(x).replace('.','',1).replace('-','',1).isdigit():
                        return str(int(float(x)))
                    else:
                        return str(x)
                except:
                    return ''
            
            df_ofimatic['idOrder_mapeado'] = df_ofimatic['idOrder_mapeado'].apply(limpiar_idorder)
            
            # Construir ID Referencia final: DOCUMENTO_ASOCIADO-NUMERO_DE_PEDIDO
            # Solo si idOrder_mapeado tiene un valor válido (no vacío, no nan)
            # Ejemplo: BG323213-089823912
            def construir_id_referencia(row):
                id_order = row['idOrder_mapeado']
                # Verificar que idOrder_mapeado no esté vacío y no sea 'nan'
                if id_order and str(id_order).strip() != '' and str(id_order).lower() != 'nan':
                    return f"{row['documento_asociado_mapeado']}-{id_order}"
                else:
                    # Si no hay NUMERO DE PEDIDO válido, dejar solo el Nrodcto original
                    return row['Nrodcto']
            
            df_ofimatic['Nrodcto_relacionado'] = df_ofimatic.apply(construir_id_referencia, axis=1)
            
            print(f"✅ Relacionados: {(df_ofimatic['idOrder_mapeado'] != '').sum()} de {len(df_ofimatic)} registros")
            
            # Paso 2: Transformar al formato Libro2.xlsx
            print("🔄 Paso 2: Transformando al formato Libro2...")
            
            # Crear DataFrame con estructura de Libro2
            df_libro2 = pd.DataFrame()
            
            # Nombre Vehiculo = NomMensajero (de ofimatic)
            df_libro2['Nombre Vehiculo'] = df_ofimatic['NomMensajero'] if 'NomMensajero' in df_ofimatic.columns else ''
            
            # Título de la Visita = NOMBRE (de ofimatic)
            df_libro2['Título de la Visita'] = df_ofimatic['NOMBRE'] if 'NOMBRE' in df_ofimatic.columns else ''
            
            # Dirección = Dirección de ehlpharma + ", " + ciudad + ", Cundinamarca"
            df_libro2['Dirección'] = df_ofimatic.apply(
                lambda row: self._construir_direccion_bogota(row),
                axis=1
            )
            
            # Latitud y Longitud - vacíos
            df_libro2['Latitud'] = None
            df_libro2['Longitud'] = None
            
            # ID Referencia = Nrodcto relacionado
            df_libro2['ID Referencia'] = df_ofimatic['Nrodcto_relacionado']
            
            # Notas = TipoVta (de ofimatic)
            df_libro2['Notas'] = df_ofimatic['TipoVta'] if 'TipoVta' in df_ofimatic.columns else ''
            
            # Persona de Contacto - vacío
            df_libro2['Persona de Contacto'] = None
            
            # Teléfono = De ehlpharma si existe, sino TEL1 o TEL2 de ofimatic
            df_libro2['Teléfono'] = df_ofimatic.apply(
                lambda row: self._obtener_telefono(row),
                axis=1
            )
            
            # Emails - vacío
            df_libro2['Emails'] = None
            
            print(f"✅ DataFrame Libro2 creado: {len(df_libro2)} registros")
            
            # Generar archivo Excel
            print("💾 Generando archivo Excel formato Libro2...")
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df_libro2.to_excel(writer, sheet_name='Hoja1', index=False)
                
                # Ajustar anchos de columna
                workbook = writer.book
                worksheet = writer.sheets['Hoja1']
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            excel_buffer.seek(0)
            excel_data = base64.b64encode(excel_buffer.read()).decode('utf-8')
            
            print(f"✅ Proceso completado: {len(df_libro2)} registros en formato Libro2")
            
            from datetime import datetime
            fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            return {
                'success': True,
                'message': f'Archivo transformado exitosamente. {len(df_libro2)} registros en formato Libro2 Bogotá.',
                'excel_data': excel_data,
                'filename': f'Libro2_Bogota_{fecha_actual}.xlsx'
            }
            
        except Exception as e:
            print(f"❌ Error: {e}")
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'error': f'Error al transformar a Libro2 Bogotá: {str(e)}',
                'details': 'Verifica que los archivos tengan las columnas correctas'
            }
    
    def _normalizar_documento_asociado(self, documento):
        """
        Normaliza el DOCUMENTO ASOCIADO de ehlpharma para facilitar la relación:
        - "bg-753083" → "BG753083"
        - "BG-753083" → "BG753083"
        - "BG753083" → "BG753083"
        - Convierte a MAYÚSCULAS y elimina guiones
        """
        if not documento or pd.isna(documento):
            return ''
        
        documento_str = str(documento).strip().upper()
        # Eliminar guiones
        documento_str = documento_str.replace('-', '')
        
        return documento_str
    
    def _normalizar_ciudad(self, ciudad):
        """
        Normaliza el nombre de la ciudad:
        - Elimina tildes y caracteres especiales
        - Convierte a MAYÚSCULAS
        """
        if not ciudad or pd.isna(ciudad):
            return ''
        
        ciudad_str = str(ciudad).strip()
        
        # Mapeo de caracteres con tilde a sin tilde
        tildes = {
            'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
            'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U',
            'ñ': 'n', 'Ñ': 'N',
            'ü': 'u', 'Ü': 'U'
        }
        
        # Reemplazar caracteres con tilde
        for con_tilde, sin_tilde in tildes.items():
            ciudad_str = ciudad_str.replace(con_tilde, sin_tilde)
        
        # Convertir a mayúsculas
        return ciudad_str.upper()
    
    def _construir_direccion(self, row, mapeo_address, mapeo_city):
        """
        Construye la dirección en formato: direccion, ciudad, Antioquia
        Prioriza la dirección de la planilla madre, si no existe usa la de ofimatic
        """
        nit = str(row['nit'])
        ciudad_raw = mapeo_city.get(nit, row.get('Destino', ''))
        
        # Normalizar ciudad (sin tildes, en MAYÚSCULAS)
        ciudad = self._normalizar_ciudad(ciudad_raw)
        
        # Prioridad 1: Dirección de planilla madre
        if nit in mapeo_address and mapeo_address[nit]:
            direccion_base = str(mapeo_address[nit]).strip()
        # Prioridad 2: Dirección de planilla ofimatic
        elif 'DIRECCION' in row and pd.notna(row['DIRECCION']):
            direccion_base = str(row['DIRECCION']).strip()
        else:
            direccion_base = ''
        
        # Agregar coma al final de la dirección si no está vacía
        if direccion_base and not direccion_base.endswith(','):
            direccion_base += ','
        
        # Construir dirección completa: "direccion, ciudad, Antioquia"
        if direccion_base and ciudad:
            return f"{direccion_base} {ciudad}, Antioquia"
        elif direccion_base:
            return f"{direccion_base} Antioquia"
        elif ciudad:
            return f"{ciudad}, Antioquia"
        else:
            return "Antioquia"
    
    def _obtener_telefono_medellin(self, row):
        """
        Obtiene el teléfono para Medellín priorizando el de la planilla madre (phonePatient_madre)
        Limpia decimales (.0) y valida que no sean números inválidos como 000
        Si TEL1 es inválido, prueba con TEL2 antes de descartar
        """
        def limpiar_telefono(telefono):
            """Limpia y valida el teléfono"""
            if pd.isna(telefono):
                return None
            
            telefono_str = str(telefono).strip()
            
            # Eliminar .0 al final si existe
            if telefono_str.endswith('.0'):
                telefono_str = telefono_str[:-2]
            
            # Convertir a int y luego a str si es un número flotante
            try:
                if '.' in str(telefono):
                    telefono_str = str(int(float(telefono)))
            except:
                pass
            
            # Validar que no sea vacío, nan, none, o inválido
            if not telefono_str or telefono_str.lower() in ['nan', 'none', '']:
                return None
            
            # Validar que no sea solo ceros (000, 0000, etc.)
            if telefono_str.replace('0', '') == '':
                return None
            
            return telefono_str
        
        # Prioridad 1: Teléfono de planilla madre (columna phonePatient_madre)
        if 'phonePatient_madre' in row and row['phonePatient_madre']:
            telefono = limpiar_telefono(row['phonePatient_madre'])
            if telefono:
                return telefono
        
        # Prioridad 2: TEL1 de planilla ofimatic
        tel1_valido = None
        if 'TEL1' in row and pd.notna(row['TEL1']):
            tel1_valido = limpiar_telefono(row['TEL1'])
        
        # Prioridad 3: TEL2 de planilla ofimatic
        tel2_valido = None
        if 'TEL2' in row and pd.notna(row['TEL2']):
            tel2_valido = limpiar_telefono(row['TEL2'])
        
        # Si TEL1 es válido, usarlo
        if tel1_valido:
            return tel1_valido
        
        # Si TEL1 no es válido pero TEL2 sí, usar TEL2
        if tel2_valido:
            return tel2_valido
        
        return None
    
    def _obtener_telefono(self, row):
        """
        Obtiene el teléfono priorizando el de la planilla ehlpharma (desde phone_ehlpharma)
        Limpia decimales (.0) y valida que no sean números inválidos como 000
        Si TEL1 es inválido, prueba con TEL2 antes de descartar
        """
        def limpiar_telefono(telefono):
            """Limpia y valida el teléfono"""
            if pd.isna(telefono):
                return None
            
            telefono_str = str(telefono).strip()
            
            # Eliminar .0 al final si existe
            if telefono_str.endswith('.0'):
                telefono_str = telefono_str[:-2]
            
            # Convertir a int y luego a str si es un número flotante
            try:
                if '.' in str(telefono):
                    telefono_str = str(int(float(telefono)))
            except:
                pass
            
            # Validar que no sea vacío, nan, none, o inválido
            if not telefono_str or telefono_str.lower() in ['nan', 'none', '']:
                return None
            
            # Validar que no sea solo ceros (000, 0000, etc.)
            if telefono_str.replace('0', '') == '':
                return None
            
            return telefono_str
        
        # Prioridad 1: Teléfono de planilla ehlpharma (columna phone_ehlpharma)
        if 'phone_ehlpharma' in row and row['phone_ehlpharma']:
            telefono = limpiar_telefono(row['phone_ehlpharma'])
            if telefono:
                return telefono
        
        # Prioridad 2: TEL1 de planilla ofimatic
        tel1_valido = None
        if 'TEL1' in row and pd.notna(row['TEL1']):
            tel1_valido = limpiar_telefono(row['TEL1'])
        
        # Prioridad 3: TEL2 de planilla ofimatic
        tel2_valido = None
        if 'TEL2' in row and pd.notna(row['TEL2']):
            tel2_valido = limpiar_telefono(row['TEL2'])
        
        # Si TEL1 es válido, usarlo
        if tel1_valido:
            return tel1_valido
        
        # Si TEL1 no es válido pero TEL2 sí, usar TEL2
        if tel2_valido:
            return tel2_valido
        
        return None
    
    def _construir_direccion_bogota(self, row):
        """
        Construye la dirección para Bogotá en formato: direccion, ciudad, Cundinamarca
        - Si hay relación (idOrder_mapeado válido): Usa dirección de Helpharma (address_ehlpharma)
        - Si no hay relación: Usa dirección de Ofimatic
        - Extrae ciudad de "Zipaquirá-Cundinamarca-Colombia" → "ZIPAQUIRA"
        - Extrae ciudad de "B-CIUDAD" → "CIUDAD" 
        - Normaliza tildes y convierte a MAYÚSCULAS
        """
        tiene_relacion = row.get('idOrder_mapeado', '') and str(row.get('idOrder_mapeado', '')).strip() != '' and str(row.get('idOrder_mapeado', '')).lower() != 'nan'
        
        # Obtener ciudad de ehlpharma (CIUDAD DE ENTREGA) - columna city_ehlpharma
        ciudad_ehlpharma = row.get('city_ehlpharma', '')
        
        # Obtener ciudad de ofimatic (Destino con formato "B-CIUDAD")
        ciudad_ofimatic = row.get('Destino', '')
        
        # Procesar ciudad
        ciudad = self._extraer_ciudad_bogota(ciudad_ehlpharma, ciudad_ofimatic)
        
        # Prioridad 1: Si hay relación, usar dirección de Helpharma (address_ehlpharma)
        if tiene_relacion and 'address_ehlpharma' in row and row['address_ehlpharma'] and str(row['address_ehlpharma']).strip() != '':
            direccion_base = str(row['address_ehlpharma']).strip()
        # Prioridad 2: Si no hay relación, usar dirección de Ofimatic
        elif 'DIRECCION' in row and pd.notna(row['DIRECCION']) and str(row['DIRECCION']).strip() != '':
            direccion_base = str(row['DIRECCION']).strip()
        # Prioridad 3: Si tampoco está en Ofimatic pero hay dirección en Helpharma, usarla
        elif 'address_ehlpharma' in row and row['address_ehlpharma'] and str(row['address_ehlpharma']).strip() != '':
            direccion_base = str(row['address_ehlpharma']).strip()
        else:
            direccion_base = ''
        
        # Agregar coma al final de la dirección si no está vacía
        if direccion_base and not direccion_base.endswith(','):
            direccion_base += ','
        
        # Construir dirección completa: "direccion, ciudad, Cundinamarca"
        if direccion_base and ciudad:
            return f"{direccion_base} {ciudad}, Cundinamarca"
        elif direccion_base:
            return f"{direccion_base} Cundinamarca"
        elif ciudad:
            return f"{ciudad}, Cundinamarca"
        else:
            return "Cundinamarca"
    
    def _extraer_ciudad_bogota(self, ciudad_ehlpharma, ciudad_ofimatic):
        """
        Extrae y normaliza el nombre de la ciudad para Bogotá
        
        Casos:
        - "Zipaquirá-Cundinamarca-Colombia" → "ZIPAQUIRA"
        - "ZIPAQUIRÁ" → "ZIPAQUIRA"
        - "BOGOTÁ. D.C." → "BOGOTA"
        - "B-BOGOTA" → "BOGOTA"
        - "B-SOACHA" → "SOACHA"
        """
        ciudad = ''
        
        # Prioridad 1: Ciudad de ehlpharma
        if ciudad_ehlpharma and pd.notna(ciudad_ehlpharma):
            ciudad_str = str(ciudad_ehlpharma).strip()
            
            # Caso: "Zipaquirá-Cundinamarca-Colombia"
            if '-' in ciudad_str:
                ciudad = ciudad_str.split('-')[0].strip()
            # Caso: "BOGOTÁ. D.C." o "ZIPAQUIRÁ"
            else:
                # Eliminar ". D.C." si existe
                ciudad = ciudad_str.replace('. D.C.', '').replace('.D.C.', '').strip()
        
        # Prioridad 2: Ciudad de ofimatic (formato "B-CIUDAD")
        if not ciudad and ciudad_ofimatic and pd.notna(ciudad_ofimatic):
            ciudad_str = str(ciudad_ofimatic).strip()
            
            # Caso: "B-BOGOTA", "B-SOACHA"
            if ciudad_str.startswith('B-'):
                ciudad = ciudad_str[2:].strip()  # Quitar "B-"
            else:
                ciudad = ciudad_str
        
        # Normalizar: sin tildes y en MAYÚSCULAS
        return self._normalizar_ciudad(ciudad)
    
    def send_json_response(self, data):
        json_data = json.dumps(data, ensure_ascii=False)
        self.send_response(200)
        self.send_header('Content-type', 'application/json; charset=utf-8')
        self.send_header('Content-length', len(json_data.encode('utf-8')))
        self.end_headers()
        self.wfile.write(json_data.encode('utf-8'))

def start_server(port=8080):
    """Inicia el servidor web local o en Render"""
    try:
        # Obtener puerto de variables de entorno (Render lo proporciona)
        port = int(os.environ.get('PORT', port))
        
        # Determinar si estamos en producción o desarrollo
        is_production = os.environ.get('RENDER', False)
        
        # En producción (Render), escuchar en 0.0.0.0, en desarrollo en localhost
        host = '0.0.0.0' if is_production else 'localhost'
        server_address = (host, port)
        
        httpd = HTTPServer(server_address, MailboxHandler)
        
        print(f"🚀 Iniciando servidor en http://{host}:{port}")
        print("📂 Directorio actual:", os.getcwd())
        print(f"🌍 Modo: {'Producción (Render)' if is_production else 'Desarrollo (Local)'}")
        print("✅ Servidor iniciado correctamente")
        
        # Solo abrir navegador en modo desarrollo (local)
        if not is_production:
            print("🌐 Abriendo navegador automáticamente...")
            
            # Abrir navegador automáticamente después de un breve retraso
            def open_browser():
                time.sleep(1.5)
                webbrowser.open(f'http://localhost:{port}')
            
            browser_thread = threading.Thread(target=open_browser)
            browser_thread.daemon = True
            browser_thread.start()
            
            print("\n" + "="*50)
            print("💡 INSTRUCCIONES:")
            print("1. El navegador se abrirá automáticamente")
            print(f"2. Si no se abre, ve a: http://localhost:{port}")
            print("3. Para cerrar: presiona Ctrl+C en esta terminal")
            print("="*50 + "\n")
        else:
            print("\n" + "="*50)
            print("✅ Servidor desplegado en Render")
            print(f"🌐 Puerto: {port}")
            print("="*50 + "\n")
        
        httpd.serve_forever()
        
    except KeyboardInterrupt:
        print("\n🛑 Cerrando servidor...")
        httpd.shutdown()
        print("👋 ¡Hasta pronto!")
    except Exception as e:
        print(f"❌ Error al iniciar el servidor: {e}")
        print("💡 Intenta con otro puerto o verifica que el puerto esté libre")

def main():
    print("🚀 Creador de Relaciones Mailbox - Versión Web")
    print("="*50)
    
    # Verificar dependencias
    try:
        import pandas
        print("✅ pandas disponible")
    except ImportError:
        print("❌ pandas no está instalado. Ejecuta: pip install pandas")
        return
    
    # Buscar puerto disponible
    for port in range(8080, 8090):
        try:
            start_server(port)
            break
        except OSError as e:
            if "Address already in use" in str(e):
                print(f"⚠️  Puerto {port} ocupado, probando {port + 1}...")
                continue
            else:
                raise

if __name__ == "__main__":
    main()