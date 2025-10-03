import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
import openpyxl

def leer_excel_inteligente(ruta_archivo):
    """
    Lee un archivo Excel detectando automáticamente dónde comienzan los datos reales
    """
    # Primero intentar leer normalmente
    try:
        df = pd.read_excel(ruta_archivo)
        # Verificar si tiene columnas conocidas
        columnas_madre = ['idOrder', 'authorizationNumber', 'typeOrder', 'identificationPatient']
        columnas_ofimatic = ['nit', 'Nrodcto']
        
        if any(col in df.columns for col in columnas_madre + columnas_ofimatic):
            return df
    except:
        pass
    
    # Si falla, buscar los encabezados usando openpyxl
    wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
    ws = wb.active
    
    # Buscar la fila que contiene los encabezados
    fila_encabezados = None
    columnas_objetivo = ['idOrder', 'authorizationNumber', 'identificationPatient', 'nit', 'Nrodcto']
    
    for fila in range(1, min(20, ws.max_row + 1)):  # Buscar en las primeras 20 filas
        valores_fila = []
        for columna in range(1, min(50, ws.max_column + 1)):  # Buscar en las primeras 50 columnas
            celda = ws.cell(row=fila, column=columna)
            if celda.value:
                valores_fila.append(str(celda.value).strip())
        
        # Verificar si esta fila contiene al menos 1 de las columnas objetivo
        coincidencias = sum(1 for col in columnas_objetivo if col in valores_fila)
        if coincidencias >= 1:
            fila_encabezados = fila - 1  # -1 porque skiprows cuenta desde 0
            break
    
    wb.close()
    
    # Leer el archivo con skiprows si encontramos los encabezados
    if fila_encabezados is not None and fila_encabezados > 0:
        df = pd.read_excel(ruta_archivo, skiprows=fila_encabezados)
    else:
        # Si no encontramos encabezados, intentar con skiprows común
        try:
            df = pd.read_excel(ruta_archivo, skiprows=4)
        except:
            df = pd.read_excel(ruta_archivo)
    
    return df

def leer_archivo_ofimatic(ruta_archivo):
    """
    Lee un archivo ofimatic (CSV o Excel) detectando automáticamente los headers
    """
    try:
        extension = os.path.splitext(ruta_archivo)[1].lower()
        
        if extension == '.csv':
            # Para CSV, intentar diferentes combinaciones
            codificaciones = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
            
            for encoding in codificaciones:
                try:
                    # Leer las primeras filas para encontrar los headers
                    df_preview = pd.read_csv(ruta_archivo, encoding=encoding, nrows=20)
                    
                    # Buscar la fila que contiene 'nit' y 'Nrodcto'
                    for skip_rows in range(10):
                        try:
                            df_test = pd.read_csv(ruta_archivo, skiprows=skip_rows, encoding=encoding, delimiter=';')
                            if 'nit' in df_test.columns and 'Nrodcto' in df_test.columns:
                                print(f"✅ Headers encontrados en fila {skip_rows + 1} con codificación {encoding}")
                                return df_test
                        except:
                            try:
                                df_test = pd.read_csv(ruta_archivo, skiprows=skip_rows, encoding=encoding, delimiter=',')
                                if 'nit' in df_test.columns and 'Nrodcto' in df_test.columns:
                                    print(f"✅ Headers encontrados en fila {skip_rows + 1} con codificación {encoding}")
                                    return df_test
                            except:
                                continue
                except:
                    continue
                    
        elif extension in ['.xlsx', '.xls']:
            # Usar la función inteligente para Excel
            df = leer_excel_inteligente(ruta_archivo)
            
            # Verificar que tenga las columnas necesarias para ofimatic
            if 'nit' not in df.columns or 'Nrodcto' not in df.columns:
                print("⚠️ Columnas 'nit' y 'Nrodcto' no encontradas, intentando detectar automáticamente...")
                
                # Si no tiene las columnas correctas, intentar detectar automáticamente
                for col_idx in range(min(15, len(df.columns))):
                    sample_data = df.iloc[:, col_idx].dropna().astype(str)
                    if len(sample_data) > 0:
                        # Si la mayoría de valores son numéricos, podría ser NIT
                        numeric_count = sum(1 for x in sample_data if x.isdigit())
                        if numeric_count > len(sample_data) * 0.7:  # 70% son números
                            # Buscar una columna cercana que pueda ser Nrodcto
                            for nrodcto_idx in range(max(0, col_idx-3), min(len(df.columns), col_idx+4)):
                                if nrodcto_idx != col_idx:
                                    sample_nrodcto = df.iloc[:, nrodcto_idx].dropna().astype(str)
                                    if len(sample_nrodcto) > 0:
                                        # Renombrar las columnas
                                        columnas_nuevas = df.columns.tolist()
                                        columnas_nuevas[col_idx] = 'nit'
                                        columnas_nuevas[nrodcto_idx] = 'Nrodcto'
                                        df.columns = columnas_nuevas
                                        print(f"🔍 Detectadas columnas: nit=columna_{col_idx}, Nrodcto=columna_{nrodcto_idx}")
                                        return df
                
                raise ValueError("No se pueden identificar las columnas 'nit' y 'Nrodcto'")
            
            return df
        
        raise Exception("No se pudo leer el archivo ofimatic")
        
    except Exception as e:
        raise Exception(f"Error al leer archivo ofimatic {ruta_archivo}: {str(e)}")

def leer_archivo(ruta_archivo):
    """
    Lee un archivo CSV o Excel y retorna un DataFrame
    """
    try:
        # Detectar la extensión del archivo
        extension = os.path.splitext(ruta_archivo)[1].lower()
        
        if extension == '.csv':
            # Lista de codificaciones a probar
            codificaciones = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252', 'utf-16']
            
            # Intentar leer CSV con diferentes delimitadores y codificaciones
            for encoding in codificaciones:
                try:
                    # Probar con punto y coma
                    df = pd.read_csv(ruta_archivo, delimiter=';', encoding=encoding)
                    print(f"✅ Archivo leído con codificación: {encoding} y delimitador ';'")
                    return df
                except:
                    try:
                        # Probar con coma
                        df = pd.read_csv(ruta_archivo, delimiter=',', encoding=encoding)
                        print(f"✅ Archivo leído con codificación: {encoding} y delimitador ','")
                        return df
                    except:
                        try:
                            # Probar con delimitador automático
                            df = pd.read_csv(ruta_archivo, encoding=encoding)
                            print(f"✅ Archivo leído con codificación: {encoding} y delimitador automático")
                            return df
                        except:
                            continue
            
            # Si todas las codificaciones fallan, intentar con errores='ignore'
            try:
                df = pd.read_csv(ruta_archivo, delimiter=';', encoding='utf-8', errors='ignore')
                print("⚠️ Archivo leído con errores ignorados")
                return df
            except:
                raise Exception("No se pudo leer el archivo CSV con ninguna codificación")
            
        elif extension in ['.xlsx', '.xls']:
            # Leer archivo Excel con función inteligente
            df = leer_excel_inteligente(ruta_archivo)
            return df
        else:
            raise ValueError(f"Formato de archivo no soportado: {extension}")
            
    except Exception as e:
        raise Exception(f"Error al leer el archivo {ruta_archivo}: {str(e)}")

# ========== FUNCIONES PARA MODO BOGOTÁ ==========

def leer_planilla_inicial_bogota(ruta_archivo):
    """
    Lee la planilla inicial de Bogotá manteniendo el formato original.
    La planilla tiene 3 filas de encabezado antes de los datos.
    """
    # Leer todo el archivo sin procesar
    df_completo = pd.read_excel(ruta_archivo, header=None)
    
    # Los encabezados están en la fila 3 (índice 3)
    encabezados = df_completo.iloc[3].tolist()
    
    # Los datos comienzan desde la fila 4 (índice 4)
    df_datos = pd.read_excel(ruta_archivo, skiprows=4)
    df_datos.columns = encabezados
    
    # Guardar las primeras 4 filas para mantener el formato original
    filas_encabezado = df_completo.iloc[0:4]
    
    return df_datos, filas_encabezado, encabezados


def leer_planilla_pedidos_bogota(ruta_archivo):
    """
    Lee la planilla de pedidos con la estructura actual.
    """
    df = pd.read_excel(ruta_archivo)
    
    # Convertir IDENTIFICACION a string para facilitar la comparación
    if 'IDENTIFICACION' in df.columns:
        df['IDENTIFICACION'] = df['IDENTIFICACION'].astype(str).str.strip()
    
    return df


def relacionar_por_nit_bogota(df_inicial, df_pedidos):
    """
    Relaciona las planillas por NIT o por DOCUMENTO ASOCIADO.
    Actualiza el campo Nrodcto con el formato: Nrodcto-NUMERO_DE_PEDIDO
    
    Intenta dos métodos de relación:
    1. Por NIT (nit == IDENTIFICACION)
    2. Por documento (Nrodcto normalizado == DOCUMENTO ASOCIADO normalizado)
    """
    # Convertir nit a string para comparación
    df_inicial['nit'] = df_inicial['nit'].astype(str).str.strip()
    
    def normalizar_documento(doc):
        """Normaliza un documento quitando guiones y convirtiendo a mayúsculas"""
        if pd.isna(doc):
            return ''
        doc_str = str(doc).strip().upper()
        # Quitar guiones y espacios
        doc_str = doc_str.replace('-', '').replace(' ', '')
        return doc_str
    
    # Crear diccionarios de mapeo
    # 1. Diccionario NIT -> NUMERO DE PEDIDO
    pedidos_por_nit = {}
    # 2. Diccionario DOCUMENTO NORMALIZADO -> NUMERO DE PEDIDO
    pedidos_por_doc = {}
    
    for _, row in df_pedidos.iterrows():
        # Convertir NUMERO DE PEDIDO a string sin decimales
        num_pedido = row['NUMERO DE PEDIDO']
        if pd.notna(num_pedido):
            try:
                num_pedido = str(int(float(num_pedido)))
            except:
                num_pedido = str(num_pedido).strip()
        else:
            num_pedido = ''
        
        # Mapeo por NIT
        nit = str(row['IDENTIFICACION']).strip()
        pedidos_por_nit[nit] = num_pedido
        
        # Mapeo por DOCUMENTO ASOCIADO
        if 'DOCUMENTO ASOCIADO' in row and pd.notna(row['DOCUMENTO ASOCIADO']):
            doc_normalizado = normalizar_documento(row['DOCUMENTO ASOCIADO'])
            if doc_normalizado:
                pedidos_por_doc[doc_normalizado] = num_pedido
    
    print(f"Total de NITs en pedidos: {len(pedidos_por_nit)}")
    print(f"Total de DOCUMENTOS en pedidos: {len(pedidos_por_doc)}")
    print(f"Total de registros en planilla inicial: {len(df_inicial)}")
    
    # Actualizar el campo Nrodcto
    registros_actualizados_nit = 0
    registros_actualizados_doc = 0
    registros_no_encontrados = []
    
    for idx, row in df_inicial.iterrows():
        nit = str(row['nit']).strip()
        nrodcto_actual = str(row['Nrodcto'])
        nrodcto_normalizado = normalizar_documento(nrodcto_actual)
        
        num_pedido = None
        metodo = None
        
        # Método 1: Intentar por NIT
        if nit in pedidos_por_nit and pedidos_por_nit[nit]:
            num_pedido = pedidos_por_nit[nit]
            metodo = 'NIT'
            registros_actualizados_nit += 1
        # Método 2: Si no encontró por NIT, intentar por DOCUMENTO
        elif nrodcto_normalizado in pedidos_por_doc and pedidos_por_doc[nrodcto_normalizado]:
            num_pedido = pedidos_por_doc[nrodcto_normalizado]
            metodo = 'DOCUMENTO'
            registros_actualizados_doc += 1
        
        # Si encontró el pedido por algún método, actualizar
        if num_pedido and metodo:
            # Crear el nuevo formato: Nrodcto-NUMERO_DE_PEDIDO
            nuevo_nrodcto = f"{nrodcto_actual}-{num_pedido}"
            df_inicial.at[idx, 'Nrodcto'] = nuevo_nrodcto
        else:
            registros_no_encontrados.append(f"{nit}|{nrodcto_actual}")
    
    total_actualizados = registros_actualizados_nit + registros_actualizados_doc
    print(f"\nRegistros actualizados por NIT: {registros_actualizados_nit}")
    print(f"Registros actualizados por DOCUMENTO: {registros_actualizados_doc}")
    print(f"Total actualizados: {total_actualizados}")
    print(f"Registros sin coincidencia: {len(registros_no_encontrados)}")
    
    return df_inicial


def guardar_con_formato_bogota(df_datos, filas_encabezado, ruta_guardado):
    """
    Guarda el DataFrame manteniendo el formato original de la planilla inicial.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Crear un nuevo workbook
    wb = Workbook()
    ws = wb.active
    
    # Escribir las primeras 4 filas de encabezado originales
    for r_idx, row in enumerate(filas_encabezado.values, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Título
                cell.font = Font(bold=True, size=12)
            elif r_idx == 4:  # Encabezados de columnas
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Escribir los datos actualizados a partir de la fila 5
    for r_idx, row in enumerate(dataframe_to_rows(df_datos, index=False, header=False), start=5):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Guardar el archivo
    wb.save(ruta_guardado)
    print(f"\nArchivo guardado exitosamente: {ruta_guardado}")

# ========== FIN FUNCIONES MODO BOGOTÁ ==========

def seleccionar_archivo_1():
    archivo = filedialog.askopenfilename(
        title="Seleccionar primer archivo (Madre)",
        filetypes=[
            ("Archivos CSV y Excel", "*.csv;*.xlsx;*.xls"), 
            ("Archivos CSV", "*.csv"), 
            ("Archivos Excel", "*.xlsx;*.xls"),
            ("Todos los archivos", "*.*")
        ]
    )
    if archivo:
        entry_archivo1.delete(0, tk.END)
        entry_archivo1.insert(0, archivo)

def seleccionar_archivo_2():
    archivo = filedialog.askopenfilename(
        title="Seleccionar segundo archivo (Ofimatic)",
        filetypes=[
            ("Archivos CSV y Excel", "*.csv;*.xlsx;*.xls"), 
            ("Archivos CSV", "*.csv"), 
            ("Archivos Excel", "*.xlsx;*.xls"),
            ("Todos los archivos", "*.*")
        ]
    )
    if archivo:
        entry_archivo2.delete(0, tk.END)
        entry_archivo2.insert(0, archivo)

def seleccionar_carpeta_destino():
    carpeta = filedialog.askdirectory(title="Seleccionar carpeta donde guardar")
    if carpeta:
        entry_destino.delete(0, tk.END)
        entry_destino.insert(0, carpeta)

def procesar_archivos():
    ruta_madre = entry_archivo1.get()
    ruta_ofimatic = entry_archivo2.get()
    carpeta_destino = entry_destino.get()
    modo = modo_var.get()
    
    if not ruta_madre or not ruta_ofimatic or not carpeta_destino:
        messagebox.showerror("Error", "Por favor selecciona todos los archivos y la carpeta de destino")
        return
    
    try:
        # Mostrar progreso
        progress_label.config(text="Procesando archivos...")
        root.update()
        
        if modo == "bogota":
            # MODO BOGOTÁ
            from datetime import datetime
            
            # Leer planilla inicial de Bogotá
            df_inicial, filas_encabezado, encabezados = leer_planilla_inicial_bogota(ruta_madre)
            print(f"✅ Planilla inicial leída: {len(df_inicial)} filas")
            
            # Leer planilla de pedidos
            df_pedidos = leer_planilla_pedidos_bogota(ruta_ofimatic)
            print(f"✅ Planilla de pedidos leída: {len(df_pedidos)} filas")
            
            # Verificar columnas requeridas
            if 'nit' not in df_inicial.columns or 'Nrodcto' not in df_inicial.columns:
                raise ValueError("Planilla Inicial no tiene las columnas necesarias (nit, Nrodcto)")
            
            if 'IDENTIFICACION' not in df_pedidos.columns or 'NUMERO DE PEDIDO' not in df_pedidos.columns:
                raise ValueError("Planilla de Pedidos no tiene las columnas necesarias (IDENTIFICACION, NUMERO DE PEDIDO)")
            
            progress_label.config(text="Relacionando datos por NIT...")
            root.update()
            
            # Relacionar por NIT
            df_actualizado = relacionar_por_nit_bogota(df_inicial, df_pedidos)
            
            progress_label.config(text="Guardando archivo...")
            root.update()
            
            # Guardar con formato original
            fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_salida = f"Planilla_Relacionada_Bogota_{fecha_actual}.xlsx"
            ruta_guardado = os.path.join(carpeta_destino, nombre_salida)
            
            guardar_con_formato_bogota(df_actualizado, filas_encabezado, ruta_guardado)
            
            progress_label.config(text=f"¡Proceso completado! {len(df_actualizado)} registros procesados", foreground="green")
            messagebox.showinfo("Éxito", f"Archivo guardado exitosamente en:\n{ruta_guardado}\n\n{len(df_actualizado)} registros procesados")
            
        else:
            # MODO NORMAL
            # Crear nombre del archivo de salida
            nombre_salida = "relaciones_unidas.xlsx"
            ruta_guardado = os.path.join(carpeta_destino, nombre_salida)
            
            # Cargar los archivos usando las funciones específicas
            df_madre = leer_archivo(ruta_madre)
            
            # Para el archivo ofimatic, usar la función específica
            df_ofimatic = leer_archivo_ofimatic(ruta_ofimatic)
            
            progress_label.config(text="Procesando datos...")
            root.update()
            
            # Verificar que las columnas necesarias existan
            if 'identificationPatient' not in df_madre.columns:
                raise ValueError("El archivo madre debe tener la columna 'identificationPatient'")
            if 'idOrder' not in df_madre.columns:
                raise ValueError("El archivo madre debe tener la columna 'idOrder'")
            if 'nit' not in df_ofimatic.columns:
                raise ValueError("El archivo ofimatic debe tener la columna 'nit'")
            if 'Nrodcto' not in df_ofimatic.columns:
                raise ValueError("El archivo ofimatic debe tener la columna 'Nrodcto'")
            
            # Seleccionar y limpiar columnas
            df_madre_reducido = df_madre[['identificationPatient', 'idOrder']].copy()
            df_madre_reducido['identificationPatient'] = df_madre_reducido['identificationPatient'].astype(str)
            df_ofimatic['nit'] = df_ofimatic['nit'].astype(str)

            # En lugar de crear un nuevo DataFrame, vamos a editar el original
            # Primero, creamos un diccionario de mapeo nit -> idOrder
            mapeo_nit_idorder = df_madre_reducido.set_index('identificationPatient')['idOrder'].to_dict()
            
            # Ahora editamos directamente el DataFrame de ofimatic
            # Crear una nueva columna con el idOrder mapeado
            df_ofimatic['idOrder_mapeado'] = df_ofimatic['nit'].map(mapeo_nit_idorder).fillna('')
            df_ofimatic['idOrder_mapeado'] = df_ofimatic['idOrder_mapeado'].astype(str)
            
            # Limpiar idOrder_mapeado para evitar decimales
            df_ofimatic['idOrder_mapeado'] = df_ofimatic['idOrder_mapeado'].apply(
                lambda x: str(int(float(x))) if x and x.replace('.','',1).isdigit() else x
            )
            
            # Actualizar la columna Nrodcto DIRECTAMENTE en el DataFrame original
            df_ofimatic['Nrodcto'] = df_ofimatic['Nrodcto'].astype(str) + '-' + df_ofimatic['idOrder_mapeado']
            
            # Eliminar la columna temporal
            df_ofimatic = df_ofimatic.drop(columns=['idOrder_mapeado'])
            
            # El DataFrame df_ofimatic ahora tiene toda la información actualizada
            # pero conserva EXACTAMENTE el formato, orden y estructura original

            progress_label.config(text="Guardando archivo...")
            root.update()
            
            # Detectar el formato original del archivo ofimatic para conservarlo
            extension_original = os.path.splitext(ruta_ofimatic)[1].lower()
            
            if extension_original in ['.xlsx', '.xls']:
                # Si el original era Excel, guardamos en Excel
                nombre_salida = "relaciones_unidas.xlsx"
                ruta_guardado = os.path.join(carpeta_destino, nombre_salida)
                
                # Guardar preservando el formato Excel original
                with pd.ExcelWriter(ruta_guardado, engine='openpyxl') as writer:
                    df_ofimatic.to_excel(writer, sheet_name='Sheet1', index=False)
                    
                    # Aplicar filtros automáticos (igual que el original)
                    workbook = writer.book
                    worksheet = writer.sheets['Sheet1']
                    worksheet.auto_filter.ref = f"A1:{chr(65 + len(df_ofimatic.columns) - 1)}{len(df_ofimatic) + 1}"
                    
                    # Ajustar ancho de columnas
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            else:
                # Si el original era CSV, guardamos en CSV
                nombre_salida = "relaciones_unidas.csv"
                ruta_guardado = os.path.join(carpeta_destino, nombre_salida)
                df_ofimatic.to_csv(ruta_guardado, index=False, sep=';')
            
            progress_label.config(text="¡Proceso completado!", foreground="green")
            messagebox.showinfo("¡Éxito!", f"Proceso completado.\nEl archivo se guardó en:\n{ruta_guardado}")

    except Exception as e:
        progress_label.config(text="Error en el proceso")
        messagebox.showerror("¡Error!", f"Ocurrió un problema: \n{e}")

# Crear la interfaz gráfica
root = tk.Tk()
root.title("Unir Planillas - CSV y Excel")
root.geometry("600x400")

# Frame principal
frame_principal = ttk.Frame(root, padding="10")
frame_principal.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

# Título
titulo = ttk.Label(frame_principal, text="Unir Planillas CSV y Excel", font=("Arial", 16, "bold"))
titulo.grid(row=0, column=0, columnspan=3, pady=(0, 20))

# Selector de modo
ttk.Label(frame_principal, text="Modo de operación:").grid(row=1, column=0, sticky=tk.W, pady=5)
modo_var = tk.StringVar(value="normal")
modo_combo = ttk.Combobox(frame_principal, textvariable=modo_var, state="readonly", width=47)
modo_combo['values'] = ("normal", "bogota")
modo_combo.grid(row=1, column=1, padx=5, pady=5)

def cambiar_modo(event=None):
    modo = modo_var.get()
    if modo == "bogota":
        label_archivo1.config(text="Planilla Inicial Bogotá (.xlsx):")
        label_archivo2.config(text="Planilla de Pedidos (.xlsx):")
        instrucciones.config(text="""
INSTRUCCIONES MODO BOGOTÁ:
1. Selecciona la PLANILLA INICIAL BOGOTÁ (Excel)
2. Selecciona la PLANILLA DE PEDIDOS (Excel)
3. Selecciona la carpeta donde guardar el resultado
4. Haz clic en PROCESAR ARCHIVOS

El proceso relacionará por NIT y actualizará Nrodcto con formato:
Nrodcto-NUMERO_PEDIDO
""")
    else:
        label_archivo1.config(text="Archivo Madre (CSV/Excel):")
        label_archivo2.config(text="Archivo Ofimatic (CSV/Excel):")
        instrucciones.config(text="""
INSTRUCCIONES:
1. Selecciona el archivo MADRE (CSV o Excel)
2. Selecciona el archivo OFIMATIC (CSV o Excel)
3. Selecciona la carpeta donde guardar el resultado
4. Haz clic en PROCESAR ARCHIVOS

Formatos soportados: .csv, .xlsx, .xls
El archivo resultado conservará el formato original (CSV o Excel)
""")

modo_combo.bind('<<ComboboxSelected>>', cambiar_modo)

# Archivo 1
label_archivo1 = ttk.Label(frame_principal, text="Archivo Madre (CSV/Excel):")
label_archivo1.grid(row=2, column=0, sticky=tk.W, pady=5)
entry_archivo1 = ttk.Entry(frame_principal, width=50)
entry_archivo1.grid(row=2, column=1, padx=5, pady=5)
ttk.Button(frame_principal, text="Seleccionar", command=seleccionar_archivo_1).grid(row=2, column=2, padx=5, pady=5)

# Archivo 2
label_archivo2 = ttk.Label(frame_principal, text="Archivo Ofimatic (CSV/Excel):")
label_archivo2.grid(row=3, column=0, sticky=tk.W, pady=5)
entry_archivo2 = ttk.Entry(frame_principal, width=50)
entry_archivo2.grid(row=3, column=1, padx=5, pady=5)
ttk.Button(frame_principal, text="Seleccionar", command=seleccionar_archivo_2).grid(row=3, column=2, padx=5, pady=5)

# Carpeta destino
ttk.Label(frame_principal, text="Carpeta de destino:").grid(row=4, column=0, sticky=tk.W, pady=5)
entry_destino = ttk.Entry(frame_principal, width=50)
entry_destino.grid(row=4, column=1, padx=5, pady=5)
ttk.Button(frame_principal, text="Seleccionar", command=seleccionar_carpeta_destino).grid(row=4, column=2, padx=5, pady=5)

# Botón procesar
ttk.Button(frame_principal, text="PROCESAR ARCHIVOS", command=procesar_archivos).grid(row=5, column=0, columnspan=3, pady=20)

# Etiqueta de progreso
progress_label = ttk.Label(frame_principal, text="Listo para procesar archivos", foreground="green")
progress_label.grid(row=6, column=0, columnspan=3, pady=10)

# Instrucciones
instrucciones_text = """
INSTRUCCIONES:
1. Selecciona el archivo MADRE (CSV o Excel)
2. Selecciona el archivo OFIMATIC (CSV o Excel)
3. Selecciona la carpeta donde guardar el resultado
4. Haz clic en PROCESAR ARCHIVOS

Formatos soportados: .csv, .xlsx, .xls
El archivo resultado conservará el formato original (CSV o Excel)
"""

instrucciones = ttk.Label(frame_principal, text=instrucciones_text, justify=tk.LEFT, foreground="gray")
instrucciones.grid(row=7, column=0, columnspan=3, pady=20, sticky=tk.W)

# Configurar pesos para redimensionamiento
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)
frame_principal.columnconfigure(1, weight=1)

if __name__ == "__main__":
    root.mainloop()