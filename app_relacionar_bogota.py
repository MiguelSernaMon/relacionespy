"""
Aplicación para relacionar planillas de Bogotá
Relaciona pedidos basados en NIT y actualiza el campo Nrodcto
Mantiene el formato de "Planillas Iniciales bogota.xlsx"
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os


def leer_planilla_inicial(archivo):
    """
    Lee la planilla inicial de Bogotá manteniendo el formato original.
    La planilla tiene 3 filas de encabezado antes de los datos.
    """
    # Leer todo el archivo sin procesar
    df_completo = pd.read_excel(archivo, header=None)
    
    # Los encabezados están en la fila 3 (índice 3)
    encabezados = df_completo.iloc[3].tolist()
    
    # Los datos comienzan desde la fila 4 (índice 4)
    df_datos = pd.read_excel(archivo, skiprows=4)
    df_datos.columns = encabezados
    
    # Guardar las primeras 3 filas para mantener el formato original
    filas_encabezado = df_completo.iloc[0:4]
    
    return df_datos, filas_encabezado, encabezados


def leer_planilla_pedidos(archivo):
    """
    Lee la planilla de pedidos con la estructura actual.
    """
    df = pd.read_excel(archivo)
    
    # Convertir IDENTIFICACION a string para facilitar la comparación
    if 'IDENTIFICACION' in df.columns:
        df['IDENTIFICACION'] = df['IDENTIFICACION'].astype(str).str.strip()
    
    return df


def relacionar_por_nit(df_inicial, df_pedidos):
    """
    Relaciona las planillas por NIT o por DOCUMENTO ASOCIADO.
    Actualiza el campo Nrodcto con el formato: Nrodcto-NUMERO_DE_PEDIDO
    
    Intenta dos métodos de relación:
    1. Por NIT (nit == IDENTIFICACION)
    2. Por documento (Nrodcto normalizado == DOCUMENTO ASOCIADO normalizado)
    """
    # Convertir nit a string para comparación
    df_inicial['nit'] = df_inicial['nit'].astype(str).str.strip()
    
    def normalizar_documento(doc):
        """Normaliza un documento quitando guiones y convirtiendo a mayúsculas"""
        if pd.isna(doc):
            return ''
        doc_str = str(doc).strip().upper()
        # Quitar guiones y espacios
        doc_str = doc_str.replace('-', '').replace(' ', '')
        return doc_str
    
    # Crear diccionarios de mapeo
    # 1. Diccionario NIT -> NUMERO DE PEDIDO
    pedidos_por_nit = {}
    # 2. Diccionario DOCUMENTO NORMALIZADO -> NUMERO DE PEDIDO
    pedidos_por_doc = {}
    
    for _, row in df_pedidos.iterrows():
        # Convertir NUMERO DE PEDIDO a string sin decimales
        num_pedido = row['NUMERO DE PEDIDO']
        if pd.notna(num_pedido):
            try:
                num_pedido = str(int(float(num_pedido)))
            except:
                num_pedido = str(num_pedido).strip()
        else:
            num_pedido = ''
        
        # Mapeo por NIT
        nit = str(row['IDENTIFICACION']).strip()
        pedidos_por_nit[nit] = num_pedido
        
        # Mapeo por DOCUMENTO ASOCIADO
        if 'DOCUMENTO ASOCIADO' in row and pd.notna(row['DOCUMENTO ASOCIADO']):
            doc_normalizado = normalizar_documento(row['DOCUMENTO ASOCIADO'])
            if doc_normalizado:
                pedidos_por_doc[doc_normalizado] = num_pedido
    
    print(f"Total de NITs en pedidos: {len(pedidos_por_nit)}")
    print(f"Total de DOCUMENTOS en pedidos: {len(pedidos_por_doc)}")
    print(f"Total de registros en planilla inicial: {len(df_inicial)}")
    
    # Actualizar el campo Nrodcto
    registros_actualizados_nit = 0
    registros_actualizados_doc = 0
    registros_no_encontrados = []
    
    for idx, row in df_inicial.iterrows():
        nit = str(row['nit']).strip()
        nrodcto_actual = str(row['Nrodcto'])
        nrodcto_normalizado = normalizar_documento(nrodcto_actual)
        
        num_pedido = None
        metodo = None
        
        # Método 1: Intentar por NIT
        if nit in pedidos_por_nit and pedidos_por_nit[nit]:
            num_pedido = pedidos_por_nit[nit]
            metodo = 'NIT'
            registros_actualizados_nit += 1
        # Método 2: Si no encontró por NIT, intentar por DOCUMENTO
        elif nrodcto_normalizado in pedidos_por_doc and pedidos_por_doc[nrodcto_normalizado]:
            num_pedido = pedidos_por_doc[nrodcto_normalizado]
            metodo = 'DOCUMENTO'
            registros_actualizados_doc += 1
        
        # Si encontró el pedido por algún método, actualizar
        if num_pedido and metodo:
            # Crear el nuevo formato: Nrodcto-NUMERO_DE_PEDIDO
            nuevo_nrodcto = f"{nrodcto_actual}-{num_pedido}"
            df_inicial.at[idx, 'Nrodcto'] = nuevo_nrodcto
        else:
            registros_no_encontrados.append(f"{nit}|{nrodcto_actual}")
    
    total_actualizados = registros_actualizados_nit + registros_actualizados_doc
    print(f"\nRegistros actualizados por NIT: {registros_actualizados_nit}")
    print(f"Registros actualizados por DOCUMENTO: {registros_actualizados_doc}")
    print(f"Total actualizados: {total_actualizados}")
    print(f"Registros sin coincidencia: {len(registros_no_encontrados)}")
    
    if registros_no_encontrados and len(registros_no_encontrados) <= 5:
        print(f"Ejemplos de registros no encontrados: {registros_no_encontrados[:5]}")
    
    return df_inicial


def guardar_con_formato(df_datos, filas_encabezado, archivo_salida):
    """
    Guarda el DataFrame manteniendo el formato original de la planilla inicial.
    """
    # Crear un archivo temporal con pandas
    with pd.ExcelWriter(archivo_salida, engine='openpyxl') as writer:
        # Primero, escribir las filas de encabezado originales
        filas_encabezado.to_excel(writer, index=False, header=False, startrow=0)
        
        # Luego escribir los datos actualizados
        df_datos.to_excel(writer, index=False, header=False, startrow=4)
    
    # Ahora aplicar estilos similares al original
    wb = load_workbook(archivo_salida)
    ws = wb.active
    
    # Aplicar formato al título (fila 1)
    ws['A1'].font = Font(bold=True, size=12)
    
    # Aplicar formato a los encabezados de columnas (fila 4)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_font = Font(bold=True)
    
    for col in range(1, len(df_datos.columns) + 1):
        cell = ws.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Guardar el archivo final
    wb.save(archivo_salida)
    print(f"\nArchivo guardado exitosamente: {archivo_salida}")


def main():
    """
    Función principal que ejecuta el proceso de relación de planillas.
    """
    print("=" * 60)
    print("RELACIONADOR DE PLANILLAS DE BOGOTÁ")
    print("=" * 60)
    
    # Archivos de entrada
    archivo_inicial = "Planillas Iniciales bogota.xlsx"
    archivo_pedidos = "Planilla 01-10-2025 (1).xlsx"
    
    # Verificar que los archivos existen
    if not os.path.exists(archivo_inicial):
        print(f"ERROR: No se encuentra el archivo {archivo_inicial}")
        return
    
    if not os.path.exists(archivo_pedidos):
        print(f"ERROR: No se encuentra el archivo {archivo_pedidos}")
        return
    
    print(f"\nLeyendo {archivo_inicial}...")
    df_inicial, filas_encabezado, encabezados = leer_planilla_inicial(archivo_inicial)
    
    print(f"Leyendo {archivo_pedidos}...")
    df_pedidos = leer_planilla_pedidos(archivo_pedidos)
    
    print("\nRelacionando datos por NIT...")
    df_actualizado = relacionar_por_nit(df_inicial, df_pedidos)
    
    # Generar nombre del archivo de salida con fecha actual
    fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
    archivo_salida = f"Planilla_Relacionada_Bogota_{fecha_actual}.xlsx"
    
    print(f"\nGuardando archivo de salida...")
    guardar_con_formato(df_actualizado, filas_encabezado, archivo_salida)
    
    print("\n" + "=" * 60)
    print("PROCESO COMPLETADO")
    print("=" * 60)
    print(f"\nArchivo generado: {archivo_salida}")
    print(f"Total de registros: {len(df_actualizado)}")


if __name__ == "__main__":
    main()
