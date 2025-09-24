#!/usr/bin/env python3
"""
Script de prueba para verificar la función de lectura inteligente de Excel
"""
import pandas as pd
import openpyxl
import os

def leer_excel_inteligente(ruta_archivo):
    """
    Lee un archivo Excel detectando automáticamente dónde comienzan los datos reales
    """
    # Primero intentar leer normalmente
    try:
        df = pd.read_excel(ruta_archivo)
        # Verificar si tiene columnas conocidas
        columnas_esperadas = ['idOrder', 'authorizationNumber', 'typeOrder', 'identificationPatient', 'nit']
        if any(col in df.columns for col in columnas_esperadas):
            print(f"✅ Lectura normal exitosa. Columnas encontradas: {df.columns.tolist()[:5]}...")
            return df
    except Exception as e:
        print(f"❌ Lectura normal falló: {e}")
    
    print("🔍 Buscando encabezados en el archivo...")
    
    # Si falla o no tiene las columnas esperadas, buscar los encabezados
    wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
    ws = wb.active
    
    # Buscar la fila que contiene los encabezados
    fila_encabezados = None
    columnas_objetivo = ['idOrder', 'authorizationNumber', 'typeOrder', 'identificationPatient', 'nit']
    
    for fila in range(1, min(20, ws.max_row + 1)):  # Buscar en las primeras 20 filas
        valores_fila = []
        for columna in range(1, min(50, ws.max_column + 1)):  # Buscar en las primeras 50 columnas
            celda = ws.cell(row=fila, column=columna)
            if celda.value:
                valores_fila.append(str(celda.value).strip())
        
        print(f"Fila {fila}: {valores_fila[:10]}...")  # Mostrar primeros 10 valores
        
        # Verificar si esta fila contiene al menos 2 de las columnas objetivo
        coincidencias = sum(1 for col in columnas_objetivo if col in valores_fila)
        if coincidencias >= 2:
            fila_encabezados = fila - 1  # -1 porque skiprows cuenta desde 0
            print(f"✅ Encabezados encontrados en fila {fila} (skiprows={fila_encabezados})")
            break
    
    wb.close()
    
    # Leer el archivo con skiprows si encontramos los encabezados
    if fila_encabezados is not None and fila_encabezados > 0:
        df = pd.read_excel(ruta_archivo, skiprows=fila_encabezados)
        print(f"✅ Lectura con skiprows={fila_encabezados} exitosa")
    else:
        # Si no encontramos encabezados, intentar con skiprows común
        print("⚠️ No se encontraron encabezados específicos, intentando con skiprows=4")
        try:
            df = pd.read_excel(ruta_archivo, skiprows=4)
            print("✅ Lectura con skiprows=4 exitosa")
        except:
            print("⚠️ Fallback a lectura normal")
            df = pd.read_excel(ruta_archivo)
    
    print(f"📊 Resultado final: {df.shape[0]} filas, {df.shape[1]} columnas")
    print(f"📋 Columnas: {df.columns.tolist()}")
    
    return df

if __name__ == "__main__":
    print("🧪 Test de lectura inteligente de Excel")
    print("Este script está listo para probar archivos Excel.")
    print("Para usar: python test_lectura.py")